#!/usr/bin/env python3
"""EXHAUSTIVE pull: every plant (all pages), every plant scanned for chemical
permits + customs exemptions (threaded), production-lines for all task-linked
plants and every line type. Emits complete schema + seed workbooks. Only
submit-inspection stays documented (it is a write and is never executed)."""
import json, re, threading
import requests
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE = "https://sectorial-be.senaei.app"
EMAIL, PASSWORD = "inspector@ip.com", "PyexdjnBEa.7LYd"
HDR = {"Accept-Language": "en"}
OUT = "/Users/sikanderahmad/Documents/mim/inspection/Inspection/tmp/senaei-discovery"
DL = "/Users/sikanderahmad/Downloads"
WORKERS = 10

_local = threading.local()
def sess():
    if not getattr(_local, "s", None): _local.s = requests.Session()
    return _local.s

# login (module session)
TOKEN = requests.post(BASE + "/api/inspection/login", headers=HDR,
                      data={"email": EMAIL, "password": PASSWORD}, timeout=30).json()["token"]
AUTH = {**HDR, "Authorization": "Bearer " + TOKEN}
print("login OK")

def get(path, params=None):
    try:
        r = sess().get(BASE + path, headers=AUTH, params=params, timeout=40)
        try: return r.json()
        except Exception: return None
    except Exception:
        return None

buckets = {}
lock = threading.Lock()
def collect(entity, records):
    if not records: return
    with lock:
        buckets.setdefault(entity, []).extend([r for r in records if isinstance(r, dict)])

# ---- profile / regulations (+detail) / tasks (+detail) ----
pj = get("/api/inspection/profile")
collect("users_profile", [pj.get("user")] if isinstance(pj, dict) and pj.get("user") else [])
rj = get("/api/inspection/regulations")
for r in (rj.get("data", []) if isinstance(rj, dict) else []):
    d = get(f"/api/inspection/regulations/{r.get('id')}")
    collect("regulations", [d.get("data") if isinstance(d, dict) and d.get("data") else r])
tj = get("/api/inspection/tasks")
tasks = tj.get("data", []) if isinstance(tj, dict) else []
collect("tasks", tasks)
task_details = []
for t in tasks:
    d = get(f"/api/inspection/tasks/{t.get('id')}")
    if isinstance(d, dict) and d.get("data"):
        task_details.append(d["data"]); collect("tasks_detail", [d["data"]])
print(f"profile/regulations/tasks done (tasks={len(tasks)})")

# ---- ALL plants (paginate to exhaustion) ----
plant_numbers, plant_ids = [], []
offset = 0
while True:
    plj = get("/api/v3/inspection/plants", params={"limit": 100, "offset": offset})
    data = plj.get("data", []) if isinstance(plj, dict) else []
    if not data: break
    collect("plants", data)
    plant_numbers += [p["plant_number"] for p in data if p.get("plant_number")]
    plant_ids += [p["id"] for p in data if p.get("id")]
    offset += 100
    if len(data) < 100: break
print(f"plants: {len(buckets.get('plants', []))}")

# ---- threaded scan: permits + exemptions for every plant ----
def scan_permit(pn):
    j = get(f"/api/v3/inspection/plants/{pn}/chemical-permits", params={"limit": 50})
    collect("chemical_permits", (j or {}).get("data", []) if isinstance(j, dict) else [])
def scan_exempt(pn):
    j = get(f"/api/v3/inspection/plants/{pn}/customs-exemptions", params={"limit": 50})
    collect("customs_exemptions", (j or {}).get("data", []) if isinstance(j, dict) else [])
with ThreadPoolExecutor(max_workers=WORKERS) as ex:
    list(ex.map(scan_permit, plant_numbers))
    print(f"chemical_permits: {len(buckets.get('chemical_permits', []))}")
    list(ex.map(scan_exempt, plant_numbers))
    print(f"customs_exemptions: {len(buckets.get('customs_exemptions', []))}")

# ---- production-line: task-linked plant ULIDs, every type ----
task_plant_ids = list(dict.fromkeys(re.findall(r'"(01[0-9a-z]{24})"', json.dumps(task_details))))
LINE_TYPES = ["product", "raw_material", "tool", "machine", "end_product"]
for pid in task_plant_ids:
    for ty in LINE_TYPES:
        j = get("/api/inspection/tasks/production-line", params={"plant_id": pid, "type": ty})
        rows = (j or {}).get("data", []) if isinstance(j, dict) else []
        for row in rows:
            if isinstance(row, dict): row.setdefault("_line_type", ty)
        collect("plant_production_line", rows)
print(f"plant_production_line: {len(buckets.get('plant_production_line', []))}")

# ---- documented fallbacks (only if still empty) + submit write shape ----
op = json.load(open(f"{DL}/inspection-api.json"))
def op_ex(p): return op["paths"][p]["get"]["responses"]["200"]["content"]["application/json"]["schema"]["example"]["data"]
pm = json.load(open(f"{DL}/Inspection_API_documentation.json"))
def _find(items, name):
    for it in items:
        if "item" in it:
            r = _find(it["item"], name)
            if r: return r
        elif it["name"] == name: return it
sources = {e: ("live" if buckets.get(e) else "documented") for e in
           ["users_profile","regulations","tasks","tasks_detail","plants",
            "chemical_permits","customs_exemptions","plant_production_line"]}
if not buckets.get("chemical_permits"): collect("chemical_permits", op_ex("/api/v3/inspection/plants/{plant_number}/chemical-permits"))
if not buckets.get("customs_exemptions"): collect("customs_exemptions", op_ex("/api/v3/inspection/plants/{plant_number}/customs-exemptions"))
if not buckets.get("plant_production_line"): collect("plant_production_line", json.loads(_find(pm["item"], "Plant Production Line")["response"][0]["body"]).get("data", []))
submit_form = _find(pm["item"], "Submit Inspection")["request"]["body"]["formdata"]
sources["inspection_submission (write)"] = "documented"
json.dump(buckets, open(f"{OUT}/full_buckets.json", "w"), indent=2, ensure_ascii=False)

# ============================ flatten + workbooks ============================
def flatten(o, pre=""):
    out = {}
    for k, v in o.items():
        col = f"{pre}{k}"
        if isinstance(v, dict):
            if set(v.keys()) <= {"value", "label"}:
                out[f"{col}_value"] = v.get("value"); out[f"{col}_label"] = v.get("label")
            else: out.update(flatten(v, col + "_"))
        elif isinstance(v, list): out[col] = f"[array:{len(v)}]"
        else: out[col] = v
    return out
tables = {}
def add_rows(table, records, source, pk=None, pid=None):
    tables.setdefault(table, {"rows": [], "source": source})
    for rec in records:
        if not isinstance(rec, dict): continue
        flat = flatten(rec)
        if pk and pid is not None: flat = {pk: pid, **flat}
        tables[table]["rows"].append(flat)
        for k, v in rec.items():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                add_rows(f"{table}__{k}", v, source, f"{table}_id", rec.get("id"))
for e, recs in buckets.items(): add_rows(e, recs, sources.get(e, "live"))
add_rows("inspection_submission (write)", [{f["key"]: f.get("value") for f in submit_form}], "documented")

def itype(v):
    if v is None: return "text"
    if isinstance(v, bool): return "boolean"
    if isinstance(v, int): return "bigint"
    if isinstance(v, float): return "numeric"
    if isinstance(v, str):
        z = v.strip()
        if len(z) == 26 and z.isalnum() and z.islower(): return "text(ulid)"
        if len(z) >= 10 and z[4:5] == "-" and z[7:8] == "-": return "timestamptz" if (":" in z or "T" in z) else "date"
        return "text"
    return "jsonb"
EP = {"users_profile":"GET /profile","regulations":"GET /regulations(+/{id})","tasks":"GET /tasks",
      "tasks_detail":"GET /tasks/{id}","plant_production_line":"GET /tasks/production-line",
      "plants":"GET /v3/…/plants","chemical_permits":"GET /v3/…/chemical-permits",
      "customs_exemptions":"GET /v3/…/customs-exemptions",
      "inspection_submission (write)":"POST /tasks/submit-inspection/{task}"}
HF = Font(bold=True, color="FFFFFF"); HB = PatternFill("solid", fgColor="2F5496")

wb = Workbook(); ws = wb.active; ws.title = "schema"
ws.append(["table","column","inferred_pg_type","nullable","example_value","source","source_endpoint"])
for c in ws[1]: c.font = HF; c.fill = HB
for table, meta in tables.items():
    base = table.split("__")[0].split(" ")[0]; cols = {}
    for row in meta["rows"]:
        for c, v in row.items():
            cols.setdefault(c, {"t": set(), "n": False, "ex": None})
            cols[c]["t"].add(itype(v))
            if v is None: cols[c]["n"] = True
            elif cols[c]["ex"] is None: cols[c]["ex"] = v
    for c, m in cols.items():
        t = "text" if ("text" in m["t"] and len(m["t"]) > 1) else sorted(m["t"])[0]
        ws.append([table, c, t, "YES" if m["n"] else "NO",
                   (str(m["ex"])[:70] if m["ex"] is not None else ""), meta["source"], EP.get(base, "")])
    ws.append([])
for col, w in {"A":34,"B":40,"C":16,"D":9,"E":42,"F":12,"G":34}.items(): ws.column_dimensions[col].width = w
wb.save(f"{OUT}/senaei_schema.xlsx")

wb2 = Workbook(); wb2.remove(wb2.active)
idx = wb2.create_sheet("_INDEX"); idx.append(["table","rows","source","endpoint"])
for c in idx[1]: c.font = HF; c.fill = HB
for table, meta in tables.items():
    base = table.split("__")[0].split(" ")[0]
    idx.append([table, len(meta["rows"]), meta["source"], EP.get(base, "")])
idx.column_dimensions["A"].width = 40; idx.column_dimensions["D"].width = 34
used = set()
def sn(t):
    n = t.replace("regulations__","reg_").replace("__","_").replace(" (write)","_W")[:31]; b=n; i=1
    while n in used: i += 1; n = f"{b[:28]}_{i}"
    used.add(n); return n
for table, meta in tables.items():
    rows = meta["rows"]
    if not rows: continue
    cols = []
    for row in rows:
        for c in row:
            if c not in cols: cols.append(c)
    w = wb2.create_sheet(sn(table)); w.append(cols)
    for c in w[1]: c.font = HF; c.fill = HB
    for row in rows:
        w.append([json.dumps(row.get(c), ensure_ascii=False) if isinstance(row.get(c), (dict, list)) else row.get(c) for c in cols])
wb2.save(f"{OUT}/senaei_seed_data.xlsx")

print("\n=== FINAL (exhaustive) ===")
for t, m in tables.items():
    print(f"  {t:44s} rows={len(m['rows']):4d}  {m['source']}")
print("DONE: senaei_schema.xlsx + senaei_seed_data.xlsx + full_buckets.json")
