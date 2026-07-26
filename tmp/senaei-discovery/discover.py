#!/usr/bin/env python3
"""
Senaei API discovery — dev/demo mirror groundwork.
Logs in as inspector, hits every READ endpoint, captures responses, then emits:
  1. senaei_schema.xlsx  -> one row per (table, column, inferred_type, nullable, example)
  2. senaei_seed_data.xlsx -> one sheet (tab) per table with the actual rows
submit-inspection (POST write) is NOT executed (side-effecting); its shape is
documented separately from the Postman collection.
"""
import json, sys, datetime
import requests
from openpyxl import Workbook

BASE = "https://sectorial-be.senaei.app"
EMAIL = "inspector@ip.com"
PASSWORD = "PyexdjnBEa.7LYd"
HDR = {"Accept-Language": "en"}
OUT = "/Users/sikanderahmad/Documents/mim/inspection/Inspection/tmp/senaei-discovery"

session = requests.Session()
raw = {}   # endpoint_key -> {status, url, json}
log = []

def hit(key, method, path, token=None, params=None, data=None, note=""):
    url = BASE + path
    h = dict(HDR)
    if token: h["Authorization"] = "Bearer " + token
    try:
        r = session.request(method, url, headers=h, params=params, data=data, timeout=30)
        body = None
        try: body = r.json()
        except Exception: body = {"_non_json": r.text[:500]}
        raw[key] = {"status": r.status_code, "url": r.url, "json": body}
        log.append(f"[{r.status_code}] {method} {path} {note}")
        return body if r.status_code < 400 else None
    except Exception as e:
        raw[key] = {"status": "ERR", "url": url, "error": str(e)}
        log.append(f"[ERR] {method} {path} -> {e}")
        return None

# ---------- 1. Login ----------
login = hit("login", "POST", "/api/inspection/login",
            data={"email": EMAIL, "password": PASSWORD}, note="auth")
token = (login or {}).get("token")
if not token:
    print("LOGIN FAILED:", json.dumps(raw.get("login"), indent=2)[:800]); sys.exit(1)
print("token acquired:", token[:12], "...")

# ---------- 2. Reads ----------
hit("profile", "GET", "/api/inspection/profile", token)

regs = hit("regulations", "GET", "/api/inspection/regulations", token)
reg_id = None
if regs and isinstance(regs.get("data"), list) and regs["data"]:
    reg_id = regs["data"][0].get("id")
    if reg_id: hit("regulation_show", "GET", f"/api/inspection/regulations/{reg_id}", token)

tasks = hit("tasks", "GET", "/api/inspection/tasks", token)
task_id = None
if tasks and isinstance(tasks.get("data"), list) and tasks["data"]:
    task_id = tasks["data"][0].get("id")
    if task_id:
        hit("task_show", "GET", f"/api/inspection/tasks/{task_id}", token)

# ---------- 3. v3 public plant APIs (fetch a wide candidate set first) ----------
plants = hit("v3_plants", "GET", "/api/v3/inspection/plants", token, params={"limit": 100})
plist = (plants or {}).get("data", []) if isinstance(plants, dict) else []
# prefer 'established' plants (more likely to carry permits / lines)
plist_sorted = sorted(plist, key=lambda p: 0 if (p.get("state") or {}).get("value") == "established" else 1)
plant_ids = [p["id"] for p in plist_sorted if p.get("id")]
plant_numbers = [p["plant_number"] for p in plist_sorted if p.get("plant_number")]

def find_non_empty(base_key, candidates, path_fn=None, param_fn=None, tries=30):
    """Try each candidate until 'data' non-empty. path_fn(c)->path OR param_fn(c)->params dict."""
    for i, c in enumerate(candidates[:tries]):
        path = path_fn(c) if path_fn else "/api/inspection/tasks/production-line"
        params = param_fn(c) if param_fn else None
        r = hit(base_key, "GET", path, token, params=params, note=f"try{i+1} c={c}")
        data = (r or {}).get("data") if isinstance(r, dict) else None
        if data:
            log.append(f"   -> {base_key} NON-EMPTY at {c} ({len(data)} rows)")
            return c
    return None

# production-line: plant ULID id via query param + type
find_non_empty("production_line", plant_ids,
               param_fn=lambda c: {"plant_id": c, "type": "product"})
# chemical-permits: plant_number in path
find_non_empty("v3_chemical_permits", plant_numbers,
               path_fn=lambda c: f"/api/v3/inspection/plants/{c}/chemical-permits")
# customs-exemptions: plant_number in path
find_non_empty("v3_customs_exemptions", plant_numbers,
               path_fn=lambda c: f"/api/v3/inspection/plants/{c}/customs-exemptions")


# ---------- persist raw ----------
with open(f"{OUT}/raw_responses.json", "w") as f:
    json.dump(raw, f, indent=2, ensure_ascii=False)
print("\n".join(log))

# ============================================================================
# Flattening / schema inference
# ============================================================================
def infer_type(v):
    if v is None: return "text"          # nullable unknown -> text
    if isinstance(v, bool): return "boolean"
    if isinstance(v, int): return "bigint"
    if isinstance(v, float): return "numeric"
    if isinstance(v, str):
        s = v.strip()
        if len(s) == 26 and s.islower() and s.isalnum(): return "text(ulid)"
        # datetime-ish
        if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-": return "timestamptz" if (":" in s or "T" in s) else "date"
        return "text"
    return "jsonb"

def flatten(obj, prefix=""):
    """single object -> {col: value}; nested {value,label} -> _value/_label; other dicts recurse; lists -> child tables (handled by caller)."""
    out = {}
    for k, v in obj.items():
        col = f"{prefix}{k}"
        if isinstance(v, dict):
            # enum resource {value,label}
            if set(v.keys()) <= {"value", "label"}:
                out[f"{col}_value"] = v.get("value")
                out[f"{col}_label"] = v.get("label")
            else:
                out.update(flatten(v, col + "_"))
        elif isinstance(v, list):
            out[col] = f"[array:{len(v)}]"   # detail goes to a child table
        else:
            out[col] = v
    return out

# table_name -> list[flat row dict]
tables = {}
def add_rows(table, records, parent_key=None, parent_id=None):
    if table not in tables: tables[table] = []
    for rec in records:
        if not isinstance(rec, dict): continue
        flat = flatten(rec)
        if parent_key and parent_id is not None:
            flat = {parent_key: parent_id, **flat}
        tables[table].append(flat)
        # recurse into nested arrays as child tables
        for k, v in rec.items():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                add_rows(f"{table}__{k}", v, parent_key=f"{table}_id", parent_id=rec.get("id"))

def records_of(key):
    r = raw.get(key, {}).get("json")
    if not r: return []
    d = r.get("data", r)
    if isinstance(d, dict): return [d]
    if isinstance(d, list): return d
    return []

add_rows("users_profile", records_of("profile") or ([raw.get("profile",{}).get("json",{}).get("user")] if raw.get("profile") else []))
add_rows("regulations", records_of("regulations"))
add_rows("tasks", records_of("tasks"))
add_rows("tasks_detail", records_of("task_show"))
add_rows("plant_production_line", records_of("production_line"))
add_rows("plants", records_of("v3_plants"))
add_rows("chemical_permits", records_of("v3_chemical_permits"))
add_rows("customs_exemptions", records_of("v3_customs_exemptions"))

# ---------- Schema workbook ----------
wb = Workbook(); ws = wb.active; ws.title = "schema"
ws.append(["table", "column", "inferred_pg_type", "nullable", "example_value", "source_endpoint"])
ep_of = {"users_profile":"GET /profile","regulations":"GET /regulations","tasks":"GET /tasks",
         "plant_production_line":"GET /tasks/production-line","plants":"GET /v3/…/plants",
         "chemical_permits":"GET /v3/…/chemical-permits","customs_exemptions":"GET /v3/…/customs-exemptions"}
for table, rows in tables.items():
    cols = {}
    for row in rows:
        for c, v in row.items():
            if c not in cols: cols[c] = {"type": set(), "null": False, "ex": None}
            cols[c]["type"].add(infer_type(v))
            if v is None: cols[c]["null"] = True
            elif cols[c]["ex"] is None: cols[c]["ex"] = v
    base = table.split("__")[0]
    for c, meta in cols.items():
        t = "text" if "text" in meta["type"] and len(meta["type"]) > 1 else sorted(meta["type"])[0]
        ex = meta["ex"]
        ws.append([table, c, t, "YES" if meta["null"] else "NO",
                   (str(ex)[:60] if ex is not None else ""), ep_of.get(base, "")])
    ws.append([])
wb.save(f"{OUT}/senaei_schema.xlsx")

# ---------- Seed-data workbook (one tab per table) ----------
wb2 = Workbook(); wb2.remove(wb2.active)
def safe_sheet(name): return name[:31]
for table, rows in tables.items():
    if not rows: continue
    cols = []
    for row in rows:
        for c in row:
            if c not in cols: cols.append(c)
    ws2 = wb2.create_sheet(safe_sheet(table))
    ws2.append(cols)
    for row in rows:
        ws2.append([row.get(c) if not isinstance(row.get(c), (dict, list)) else json.dumps(row.get(c), ensure_ascii=False) for c in cols])
if not wb2.sheetnames: wb2.create_sheet("empty")
wb2.save(f"{OUT}/senaei_seed_data.xlsx")

# ---------- summary ----------
print("\n=== TABLES DISCOVERED ===")
for t, r in tables.items():
    print(f"  {t}: {len(r)} rows")
print(f"\nWrote: senaei_schema.xlsx, senaei_seed_data.xlsx, raw_responses.json  in {OUT}")
