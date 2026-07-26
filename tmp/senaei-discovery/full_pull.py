#!/usr/bin/env python3
"""FULL pull: gather as many real rows as practical from every Senaei read
endpoint, then emit the two complete workbooks (schema + seed). Bounded scans so
it finishes in a couple of minutes. submit-inspection stays documented (write)."""
import json, re, sys, time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE = "https://sectorial-be.senaei.app"
EMAIL, PASSWORD = "inspector@ip.com", "PyexdjnBEa.7LYd"
HDR = {"Accept-Language": "en"}
OUT = "/Users/sikanderahmad/Documents/mim/inspection/Inspection/tmp/senaei-discovery"
DL = "/Users/sikanderahmad/Downloads"

# scan / collection caps (keep runtime sane)
PLANT_PAGES = 3          # 100 plants per page
PERMIT_TARGET = 25       # stop once this many permit rows collected
PERMIT_SCAN = 80         # ...or after scanning this many plants
EXEMPT_TARGET, EXEMPT_SCAN = 25, 80
PRODLINE_TARGET, PRODLINE_SCAN = 40, 60

s = requests.Session()
def get(path, token=None, params=None):
    h = dict(HDR)
    if token: h["Authorization"] = "Bearer " + token
    try:
        r = s.get(BASE + path, headers=h, params=params, timeout=30)
        try: return r.status_code, r.json()
        except Exception: return r.status_code, None
    except Exception as e:
        return "ERR", {"error": str(e)}

# ---- login ----
lh = dict(HDR)
lr = s.post(BASE + "/api/inspection/login", headers=lh, data={"email": EMAIL, "password": PASSWORD}, timeout=30)
token = lr.json().get("token")
if not token: print("LOGIN FAILED", lr.text[:300]); sys.exit(1)
print("login OK, token", token[:10], "…")

buckets = {}   # entity -> list[record dict]
def collect(entity, records):
    buckets.setdefault(entity, [])
    for rec in records or []:
        if isinstance(rec, dict): buckets[entity].append(rec)

# ---- profile ----
_, pj = get("/api/inspection/profile", token)
collect("users_profile", [pj.get("user")] if isinstance(pj, dict) and pj.get("user") else [])

# ---- regulations: list + each detail ----
_, rj = get("/api/inspection/regulations", token)
regs = rj.get("data", []) if isinstance(rj, dict) else []
reg_full = []
for r in regs:
    _, d = get(f"/api/inspection/regulations/{r.get('id')}", token)
    reg_full.append(d.get("data") if isinstance(d, dict) and d.get("data") else r)
collect("regulations", reg_full)
print(f"regulations: {len(reg_full)}")

# ---- tasks: list + each detail ----
_, tj = get("/api/inspection/tasks", token)
tasks = tj.get("data", []) if isinstance(tj, dict) else []
collect("tasks", tasks)
task_details = []
for t in tasks:
    _, d = get(f"/api/inspection/tasks/{t.get('id')}", token)
    if isinstance(d, dict) and d.get("data"): task_details.append(d["data"])
collect("tasks_detail", task_details)
print(f"tasks: {len(tasks)}  details: {len(task_details)}")

# ---- plants (paginate) ----
plant_ids, plant_numbers = [], []
for pg in range(PLANT_PAGES):
    _, plj = get("/api/v3/inspection/plants", token, params={"limit": 100, "offset": pg * 100})
    data = plj.get("data", []) if isinstance(plj, dict) else []
    if not data: break
    collect("plants", data)
    for p in data:
        if p.get("id"): plant_ids.append(p["id"])
        if p.get("plant_number"): plant_numbers.append(p["plant_number"])
print(f"plants collected: {len(buckets.get('plants', []))}")

# prefer established plants first for permit/exemption scans
est = [p for p in buckets.get("plants", []) if (p.get("state") or {}).get("value") == "established"]
scan_numbers = [p["plant_number"] for p in est if p.get("plant_number")] + plant_numbers
scan_ids = [p["id"] for p in est if p.get("id")] + plant_ids
# de-dup preserving order
def dedup(x):
    seen=set(); out=[]
    for i in x:
        if i not in seen: seen.add(i); out.append(i)
    return out
scan_numbers, scan_ids = dedup(scan_numbers), dedup(scan_ids)

# ---- chemical-permits scan ----
n=0
for pn in scan_numbers[:PERMIT_SCAN]:
    _, j = get(f"/api/v3/inspection/plants/{pn}/chemical-permits", token, params={"limit": 50})
    data = j.get("data", []) if isinstance(j, dict) else []
    if data: collect("chemical_permits", data)
    n += 1
    if len(buckets.get("chemical_permits", [])) >= PERMIT_TARGET: break
print(f"chemical_permits live: {len(buckets.get('chemical_permits', []))} (scanned {n} plants)")

# ---- customs-exemptions scan ----
n=0
for pn in scan_numbers[:EXEMPT_SCAN]:
    _, j = get(f"/api/v3/inspection/plants/{pn}/customs-exemptions", token, params={"limit": 50})
    data = j.get("data", []) if isinstance(j, dict) else []
    if data: collect("customs_exemptions", data)
    n += 1
    if len(buckets.get("customs_exemptions", [])) >= EXEMPT_TARGET: break
print(f"customs_exemptions live: {len(buckets.get('customs_exemptions', []))} (scanned {n} plants)")

# ---- production-line scan ----
# IMPORTANT: production-line data is keyed on the LICENSE plant ULID embedded in
# task details (NOT the public v3 plant list). Gather those ULIDs first, then
# fall back to v3 plant ids. Try every known line type.
task_plant_ids = dedup(re.findall(r'"(01[0-9a-z]{24})"', json.dumps(task_details)))
prodline_ids = dedup(task_plant_ids + scan_ids)
LINE_TYPES = ["product", "raw_material", "tool", "machine", "end_product"]
n=0
for pid in prodline_ids[:PRODLINE_SCAN]:
    hit_any = False
    for ty in LINE_TYPES:
        _, j = get("/api/inspection/tasks/production-line", token, params={"plant_id": pid, "type": ty})
        data = j.get("data", []) if isinstance(j, dict) else []
        for row in data:
            if isinstance(row, dict): row.setdefault("_line_type", ty)
        if data: collect("plant_production_line", data); hit_any = True
    n += 1
    if len(buckets.get("plant_production_line", [])) >= PRODLINE_TARGET: break
print(f"plant_production_line live: {len(buckets.get('plant_production_line', []))} (scanned {n} plants)")

# ---- documented fallbacks for anything still empty ----
op = json.load(open(f"{DL}/inspection-api.json"))
def op_ex(path):
    return op["paths"][path]["get"]["responses"]["200"]["content"]["application/json"]["schema"]["example"]["data"]
pm = json.load(open(f"{DL}/Inspection_API_documentation.json"))
def _find(items, name):
    for it in items:
        if "item" in it:
            r = _find(it["item"], name)
            if r: return r
        elif it["name"] == name: return it
sources = {}   # entity -> 'live' | 'documented' | 'mixed'
for e in ["users_profile","regulations","tasks","tasks_detail","plants","chemical_permits","customs_exemptions","plant_production_line"]:
    sources[e] = "live" if buckets.get(e) else "documented"
if not buckets.get("chemical_permits"):
    collect("chemical_permits", op_ex("/api/v3/inspection/plants/{plant_number}/chemical-permits"))
if not buckets.get("customs_exemptions"):
    collect("customs_exemptions", op_ex("/api/v3/inspection/plants/{plant_number}/customs-exemptions"))
if not buckets.get("plant_production_line"):
    collect("plant_production_line", json.loads(_find(pm["item"], "Plant Production Line")["response"][0]["body"]).get("data", []))
# submit-inspection write shape (always documented)
submit_form = _find(pm["item"], "Submit Inspection")["request"]["body"]["formdata"]
sources["inspection_submission (write)"] = "documented"

json.dump(buckets, open(f"{OUT}/full_buckets.json","w"), indent=2, ensure_ascii=False)

# ============================================================================
# flatten + workbooks
# ============================================================================
def flatten(obj, prefix=""):
    out = {}
    for k, v in obj.items():
        col = f"{prefix}{k}"
        if isinstance(v, dict):
            if set(v.keys()) <= {"value", "label"}:
                out[f"{col}_value"] = v.get("value"); out[f"{col}_label"] = v.get("label")
            else: out.update(flatten(v, col + "_"))
        elif isinstance(v, list): out[col] = f"[array:{len(v)}]"
        else: out[col] = v
    return out

tables = {}   # table -> {"rows":[...], "source":str}
def add_rows(table, records, source, parent_key=None, parent_id=None):
    tables.setdefault(table, {"rows": [], "source": source})
    for rec in records:
        if not isinstance(rec, dict): continue
        flat = flatten(rec)
        if parent_key and parent_id is not None: flat = {parent_key: parent_id, **flat}
        tables[table]["rows"].append(flat)
        for k, v in rec.items():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                add_rows(f"{table}__{k}", v, source, f"{table}_id", rec.get("id"))

for e, recs in buckets.items():
    add_rows(e, recs, sources.get(e, "live"))
add_rows("inspection_submission (write)", [{f["key"]: f.get("value") for f in submit_form}], "documented")

def infer_type(v):
    if v is None: return "text"
    if isinstance(v, bool): return "boolean"
    if isinstance(v, int): return "bigint"
    if isinstance(v, float): return "numeric"
    if isinstance(v, str):
        z = v.strip()
        if len(z) == 26 and z.isalnum() and z.islower(): return "text(ulid)"
        if len(z) >= 10 and z[4:5] == "-" and z[7:8] == "-":
            return "timestamptz" if (":" in z or "T" in z) else "date"
        return "text"
    return "jsonb"

EP = {"users_profile":"GET /profile","regulations":"GET /regulations(+/{id})","tasks":"GET /tasks",
      "tasks_detail":"GET /tasks/{id}","plant_production_line":"GET /tasks/production-line",
      "plants":"GET /v3/…/plants","chemical_permits":"GET /v3/…/chemical-permits",
      "customs_exemptions":"GET /v3/…/customs-exemptions",
      "inspection_submission (write)":"POST /tasks/submit-inspection/{task}"}
HF = Font(bold=True, color="FFFFFF"); HB = PatternFill("solid", fgColor="2F5496")

# schema workbook
wb = Workbook(); ws = wb.active; ws.title = "schema"
ws.append(["table","column","inferred_pg_type","nullable","example_value","source","source_endpoint"])
for c in ws[1]: c.font = HF; c.fill = HB
for table, meta in tables.items():
    base = table.split("__")[0].split(" ")[0]; cols = {}
    for row in meta["rows"]:
        for c, v in row.items():
            cols.setdefault(c, {"type": set(), "null": False, "ex": None})
            cols[c]["type"].add(infer_type(v))
            if v is None: cols[c]["null"] = True
            elif cols[c]["ex"] is None: cols[c]["ex"] = v
    for c, m in cols.items():
        t = "text" if ("text" in m["type"] and len(m["type"]) > 1) else sorted(m["type"])[0]
        ws.append([table, c, t, "YES" if m["null"] else "NO",
                   (str(m["ex"])[:70] if m["ex"] is not None else ""), meta["source"], EP.get(base,"")])
    ws.append([])
for col, w in {"A":34,"B":40,"C":16,"D":9,"E":42,"F":12,"G":34}.items(): ws.column_dimensions[col].width = w
wb.save(f"{OUT}/senaei_schema.xlsx")

# seed workbook
wb2 = Workbook(); wb2.remove(wb2.active)
idx = wb2.create_sheet("_INDEX"); idx.append(["table","rows","source","endpoint"])
for c in idx[1]: c.font = HF; c.fill = HB
for table, meta in tables.items():
    base = table.split("__")[0].split(" ")[0]
    idx.append([table, len(meta["rows"]), meta["source"], EP.get(base,"")])
idx.column_dimensions["A"].width=36; idx.column_dimensions["D"].width=34
used=set()
def sn(t):
    n=t.replace("regulations__","reg_").replace("__","_").replace(" (write)","_W")[:31]; b=n; i=1
    while n in used: i+=1; n=f"{b[:28]}_{i}"
    used.add(n); return n
for table, meta in tables.items():
    rows = meta["rows"]
    if not rows: continue
    cols=[]
    for row in rows:
        for c in row:
            if c not in cols: cols.append(c)
    w = wb2.create_sheet(sn(table)); w.append(cols)
    for c in w[1]: c.font = HF; c.fill = HB
    for row in rows:
        w.append([json.dumps(row.get(c), ensure_ascii=False) if isinstance(row.get(c),(dict,list)) else row.get(c) for c in cols])
wb2.save(f"{OUT}/senaei_seed_data.xlsx")

print("\n=== FINAL ===")
for t, m in tables.items():
    print(f"  {t:44s} rows={len(m['rows']):3d}  {m['source']}")
print(f"\nWrote senaei_schema.xlsx + senaei_seed_data.xlsx + full_buckets.json")
