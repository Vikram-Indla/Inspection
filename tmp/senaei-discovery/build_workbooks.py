#!/usr/bin/env python3
"""Rebuild schema + seed workbooks from cached raw_responses.json, filling any
entity that came back EMPTY live with the documented Postman/OpenAPI example
(clearly marked source=documented). No live calls."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT = "/Users/sikanderahmad/Documents/mim/inspection/Inspection/tmp/senaei-discovery"
DL = "/Users/sikanderahmad/Downloads"
raw = json.load(open(f"{OUT}/raw_responses.json"))

# ---- fallbacks from the two source files ----
op = json.load(open(f"{DL}/inspection-api.json"))
chem_ex = op["paths"]["/api/v3/inspection/plants/{plant_number}/chemical-permits"]["get"]["responses"]["200"]["content"]["application/json"]["schema"]["example"]["data"]
custom_ex = op["paths"]["/api/v3/inspection/plants/{plant_number}/customs-exemptions"]["get"]["responses"]["200"]["content"]["application/json"]["schema"]["example"]["data"]
pm = json.load(open(f"{DL}/Inspection_API_documentation.json"))
def _find(items, name):
    for it in items:
        if "item" in it:
            r = _find(it["item"], name)
            if r: return r
        elif it["name"] == name: return it
def _pm_body(name):
    it = _find(pm["item"], name)
    return json.loads(it["response"][0]["body"])
prodline_ex = _pm_body("Plant Production Line").get("data", [])
submit_body = _find(pm["item"], "Submit Inspection")["request"]["body"]["formdata"]

def records_of(key):
    r = raw.get(key, {}).get("json")
    if not isinstance(r, (dict, list)): return []
    d = r.get("data", r) if isinstance(r, dict) else r
    if isinstance(d, dict):
        # profile wraps user
        if "user" in d and len(d) == 1: return [d["user"]]
        return [d]
    return d if isinstance(d, list) else []

# entity -> (records, source)
entities = {}
def put(name, live_key, fallback=None):
    recs = records_of(live_key)
    if recs: entities[name] = (recs, "live")
    elif fallback: entities[name] = (fallback, "documented")
    else: entities[name] = ([], "empty")

put("users_profile", "profile")
put("regulations", "regulations")
put("tasks", "tasks")
put("tasks_detail", "task_show")
put("plant_production_line", "production_line", prodline_ex)
put("plants", "v3_plants")
put("chemical_permits", "v3_chemical_permits", chem_ex)
put("customs_exemptions", "v3_customs_exemptions", custom_ex)

# ---- flatten (enum {value,label} -> _value/_label; nested arrays -> child tables) ----
def flatten(obj, prefix=""):
    out = {}
    for k, v in obj.items():
        col = f"{prefix}{k}"
        if isinstance(v, dict):
            if set(v.keys()) <= {"value", "label"}:
                out[f"{col}_value"] = v.get("value"); out[f"{col}_label"] = v.get("label")
            else:
                out.update(flatten(v, col + "_"))
        elif isinstance(v, list):
            out[col] = f"[array:{len(v)}]"
        else:
            out[col] = v
    return out

tables = {}   # table -> {"rows":[...], "source":str}
def add_rows(table, records, source, parent_key=None, parent_id=None):
    tables.setdefault(table, {"rows": [], "source": source})
    for rec in records:
        if not isinstance(rec, dict): continue
        flat = flatten(rec)
        if parent_key and parent_id is not None:
            flat = {parent_key: parent_id, **flat}
        tables[table]["rows"].append(flat)
        for k, v in rec.items():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                add_rows(f"{table}__{k}", v, source, f"{table}_id", rec.get("id"))

for name, (recs, src) in entities.items():
    add_rows(name, recs, src)

# submit-inspection -> synthesize a write-target table from the form-data keys
sub_cols = {}
for f in submit_body:
    key = f["key"]
    if "[" in key:  # array/nested -> note as child structure
        sub_cols[key] = f.get("value")
    else:
        sub_cols[key] = f.get("value")
tables["inspection_submission (write)"] = {"rows": [sub_cols], "source": "documented"}

# ---- type inference ----
def infer_type(v):
    if v is None: return "text"
    if isinstance(v, bool): return "boolean"
    if isinstance(v, int): return "bigint"
    if isinstance(v, float): return "numeric"
    if isinstance(v, str):
        s = v.strip()
        if len(s) == 26 and s.isalnum() and s.islower(): return "text(ulid)"
        if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-":
            return "timestamptz" if (":" in s or "T" in s) else "date"
        return "text"
    return "jsonb"

EP = {"users_profile":"GET /profile","regulations":"GET /regulations(+/{id})","tasks":"GET /tasks",
      "tasks_detail":"GET /tasks/{id}","plant_production_line":"GET /tasks/production-line",
      "plants":"GET /v3/…/plants","chemical_permits":"GET /v3/…/chemical-permits",
      "customs_exemptions":"GET /v3/…/customs-exemptions",
      "inspection_submission (write)":"POST /tasks/submit-inspection/{task}"}

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")

# ---- Workbook 1: schema ----
wb = Workbook(); ws = wb.active; ws.title = "schema"
ws.append(["table","column","inferred_pg_type","nullable","example_value","source","source_endpoint"])
for c in ws[1]: c.font = hdr_font; c.fill = hdr_fill
for table, meta in tables.items():
    rows = meta["rows"]; src = meta["source"]; base = table.split("__")[0].split(" ")[0]
    cols = {}
    for row in rows:
        for c, v in row.items():
            cols.setdefault(c, {"type": set(), "null": False, "ex": None})
            cols[c]["type"].add(infer_type(v))
            if v is None: cols[c]["null"] = True
            elif cols[c]["ex"] is None: cols[c]["ex"] = v
    for c, m in cols.items():
        t = "text" if ("text" in m["type"] and len(m["type"]) > 1) else sorted(m["type"])[0]
        ws.append([table, c, t, "YES" if m["null"] else "NO",
                   (str(m["ex"])[:70] if m["ex"] is not None else ""), src, EP.get(base, "")])
    ws.append([])
for col, w in {"A":34,"B":40,"C":16,"D":9,"E":40,"F":12,"G":34}.items():
    ws.column_dimensions[col].width = w
wb.save(f"{OUT}/senaei_schema.xlsx")

# ---- Workbook 2: seed data (one tab per table) ----
wb2 = Workbook(); wb2.remove(wb2.active)
used = set()
def sheet_name(t):
    n = t.replace("regulations__","reg_").replace("__","_").replace(" (write)","_WRITE")[:31]
    b = n; i = 1
    while n in used: i += 1; n = f"{b[:28]}_{i}"
    used.add(n); return n
# index tab
idx = wb2.create_sheet("_INDEX")
idx.append(["table","rows","source","endpoint"])
for c in idx[1]: c.font = hdr_font; c.fill = hdr_fill
for table, meta in tables.items():
    base = table.split("__")[0].split(" ")[0]
    idx.append([table, len(meta["rows"]), meta["source"], EP.get(base,"")])
idx.column_dimensions["A"].width=34; idx.column_dimensions["D"].width=34
for table, meta in tables.items():
    rows = meta["rows"]
    if not rows: continue
    cols = []
    for row in rows:
        for c in row:
            if c not in cols: cols.append(c)
    ws2 = wb2.create_sheet(sheet_name(table))
    ws2.append(cols)
    for c in ws2[1]: c.font = hdr_font; c.fill = hdr_fill
    for row in rows:
        ws2.append([json.dumps(row.get(c), ensure_ascii=False) if isinstance(row.get(c),(dict,list)) else row.get(c) for c in cols])
wb2.save(f"{OUT}/senaei_seed_data.xlsx")

print("=== FINAL TABLES ===")
for t, m in tables.items():
    print(f"  {t:42s} rows={len(m['rows']):3d}  source={m['source']}")
print(f"\nWrote senaei_schema.xlsx + senaei_seed_data.xlsx to {OUT}")
