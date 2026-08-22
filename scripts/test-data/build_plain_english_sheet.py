# -*- coding: utf-8 -*-
"""One sheet: As-is | To-be | Reason for user-facing English in apps/web.

Source of truth: apps/web/src/i18n/locales/en/*.json.
Roles are the four real roles only — admin, inspector, planner, supervisor —
per supabase/migrations/20260805180000_select_phantom_roles_four_real_roles.sql,
which names {ops, compliance_admin, form_admin, reviewer, auditor, leadership,
security_admin, workflow_admin} as phantom.

Run: python3 scripts/test-data/build_plain_english_sheet.py
"""
import json, glob, os, sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOCALES = os.path.join(ROOT, "apps", "web", "src", "i18n", "locales", "en")

ROLE = {
    "field": "Inspector", "planning": "Planner", "visits": "Planner",
    "reviews": "Supervisor", "approvals": "Supervisor", "tasks": "Supervisor",
    "execution": "Supervisor", "admin": "Admin",
}
def role_for(ns):
    head = ns.split("-")[0]
    return ROLE.get(head, ROLE.get(ns, "All"))

# key -> (to_be, reason)
R = {
"admin-factory-data|csv.schemaNote": ("Schema {schema}. Required columns: {required}. Files are held for review and never overwrite factory records.", "Jargon: staged, reconciliation, factory truth"),
"admin-workflows|deck.authoredElsewhere": ("Read-only. To change this, propose a draft below. Someone else must publish it.", "27 words to 13; drops payload, governed"),
"planning|bulk.notice.unreadableBody": ("That criteria link was broken. Rebuild your criteria below to see results.", "24 words to 11"),
"admin-factory-data|csv.staged": ("Batch {batch}: {pending} waiting for review, {rejected} rejected. Nothing is imported until reviewed.", "Jargon: staged, reconciliation, governed"),
"field-unregistered|note": ("Starts a visit now, assigned to you. Your photo and location identify the site until it is matched.", "28 words to 18; drops placeholder identity, reconciliation"),
"admin-audit|banners.degraded": ("Detailed event replay is off here. Basic events still show, but are not recorded as official.", "Jargon: semantic replay contracts, promoted to official facts"),
"execution|reschedule.noHandoff": ("The new day was not sent to Planning. Planning has no rule for this yet.", "Jargon: governed handoff contract"),
"admin-senai-data|subtitle": ("Read-only view of the SENAI source, its connection, and how records are matched.", "Jargon: registry, endpoint contract, source of truth"),
"enforcement|catalogue.lensRule4": ("A published mapping cannot be edited. Create a replacement instead.", "Jargon: immutable, governed successor"),
"admin-dashboard-config|catalogue.intro": ("Formula and metric ID are fixed and cannot be edited. {implemented} of {total} metrics have a working formula.", "Jargon: system-seeded, immutable to ordinary admins"),
"admin-localization|row.placeholderError": ("The Arabic is missing {token}. Add it to save.", "17 words to 8"),
"admin-planning-lookups|actions.added": ("\"{kind}.{key}\" added and is now active. It appears in planning lists after the next load.", "Jargon: audited, governed selects"),
"admin-planning-lookups|actions.deactivated": ("\"{kind}.{key}\" turned off. It stays in history but no longer appears in planning lists.", "Jargon: never deleted, governed selects"),
"admin-planning-lookups|deleteNote": ("Values are never deleted, because past records still use them. Turning one off just hides it from new entries.", "31 words to 19"),
"admin-planning-status|governance.body": ("Status changes follow the published workflow. This screen is read-only.", "27 words to 10"),
"admin-senai-data|endpoints.footer": ("Checks are based only on recorded calls. If a call fails, the linked Factory 360 section shows as unavailable rather than guessing a value.", "48 words to 24; drops table name, allow-list detail"),
"admin-senai-data|fnd007.lead": ("SENAI holds the official factory, licence and industrial records. This platform reads them and never writes back.", "Jargon: source of truth, reconciles"),
"admin-senai-data|reconcile.runsError": ("Sync history could not be loaded. Row counts and freshness are unavailable.", "Jargon: reconciliation divergences, independent"),
"admin-workflows|sla.none": ("No SLA calendar is set up. Until one is approved, no deadlines are shown.", "33 words to 14"),
"analytics|bottlenecks.body": ("Each of these needs a value that must be approved first, such as a target or a threshold.", "Jargon: cohort rule, lookback, weighting, denominator"),
"analytics|bottlenecks.exportNote": ("Waiting on an approved audit rule.", "Jargon: canonical, correlation contract"),
"analytics|empty.title": ("No data for this period", "Jargon: governed aggregates"),
"dashboard|national.coverage.interpretation": ("Needs an inspection-cycle target from Admin.", "13 words to 6; drops withheld, governed"),
"enforcement|catalogue.lensRule1": ("Each violation can have only one mapping.", "Drops internal reason: unique constraint"),
"factories-cr|government.empty": ("No government records are connected yet.", "Jargon: governed source contract"),
"field-reports|tabs.recordsHint": ("Your submitted records. Open one to see the version that was sent.", "Jargon: immutable version"),
"planning|bulk.form.selectAllConfirmBody": ("This selects all {n} matching factories, including ones on other pages. Type {n} to confirm.", "23 words to 15"),
"planning|bulkReview.droppedD": ("These rows stopped matching before submission, so none of the batch was submitted.", "Jargon: ineligible, committed atomically"),
"planning|emptyStates.noneWithTotal": ("{total} records exist, but none match this page. Some may be hidden by access rules. Refresh to try again.", "35 words to 19"),
"profile|notif.inappNote": ("The bell is always on. It is the only channel guaranteed to reach you.", "Jargon: delivery surface"),
"admin-audit|ledger.selectCase": ("Select one complete case with a published event list. Viewing all cases will not work.", "24 words to 14"),
"admin-dashboard-config|domains.intro": ("Each domain publishes a dated version that cannot be edited. Settings can narrow who sees data, never widen it.", "Jargon: immutable, effective-dated, RLS"),
"admin-factory-data|banner.body": ("Factory 360 profiles are read-only. Setting up a provider does not connect it. CSV rows wait for approval, and source IDs are never replaced silently.", "34 words to 25; drops staged, reconciliation"),
"admin-factory-data|history.reconciliation.emptyBody": ("No matching records for you. This is a real empty result, not an error.", "Jargon: staged does not mean imported"),
"admin-factory-data|history.reconciliation.error": ("Match history could not be loaded. This is an error, not an empty history.", "Jargon: reconciliation"),
"admin-factory-data|history.runsError": ("Sync history could not be loaded. Rejected rows and match records below still work.", "Jargon: reconciliation, independently readable"),
"admin-factory-data|masterData.factoriesError": ("The factory list could not be loaded. Import history below may still work.", "Jargon: reconciliation"),
"admin-factory-data|masterData.noWritebackHelp": ("Changes here are saved in this platform only. Nothing is sent back to SENAI.", "43 words to 14"),
"admin-factory-data|provider.help": ("Connection details are shown in SENAI data management. Set up is not the same as connected.", "Jargon: endpoint contract, recorded live-call check"),
"admin-integrations|banner.body": ("An endpoint is only connected once its details are approved and an address is set. Secrets are never shown.", "Jargon: contract, runtime address"),
"admin-risk|reconstruction.note": ("Weights, bands and recalculation times are not shown here. This screen reads live settings. Per-factory scores are on the factory record.", "34 words to 21"),
"admin-senai-data|endpoints.help": ("The list of SENAI calls this platform is allowed to make.", "Jargon: documented read contract, allow-list, runtime client"),
"admin-senai-data|mapping.body": ("No field mapping is set up yet, so none can be shown. Until one is approved, unmapped fields show as not configured.", "73 words to 22; drops code path, invent warning"),
"admin-senai-data|reconcile.error": ("Differences could not be loaded. This is an error, not an empty result.", "Jargon: reconciliation divergences, checked empty set"),
"admin-senai-data|reconcile.footer": ("Fixing a difference is a separate approved action, not an edit here. Nothing is overwritten or sent to SENAI.", "34 words to 19"),
"admin-templates|journey.description": ("Publish the legal source, prepare the templates, then publish the inspection package that uses them.", "Jargon: immutable, assemble, references"),
"dashboard|trend.footnote": ("Penalty notices issued this period, compared with the period before.", "18 words to 10"),
"planning|bulk.empty.noCriteria.body": ("Add a condition above and press Apply criteria. Nothing is selected by default.", "23 words to 13"),
"planning|bulk.empty.noFilterMatch.body": ("Factories matched your criteria, but none match the text you typed. Clear the filter.", "20 words to 14"),
"planning|bulk.empty.noMatch.body": ("No factories match. Widen a condition, or switch the group from ALL to ANY.", "Drops excluded by current criteria"),
"planning|bulk.form.invalidBody": ("{n} selected factories no longer match and were removed.", "16 words to 9"),
"planning|emptyStates.filteredBody": ("No visits match your search and filters. Clear them to see everything.", "Jargon: scoped"),
"profile|details.editNote": ("Your details, region and roles are managed by an administrator and cannot be changed here.", "24 words to 15; drops self-service identity"),
"regulations|authorities.caption": ("Grouped by the authority recorded on each regulation.", "Drops there is no registry to reconcile against"),
"admin-integrations|governance.points[1]": ("Secrets are never shown. Only the approved details and current state are shown.", "Jargon: contract, runtime state"),
"admin-integrations|registry.help": ("Version, current state and what depends on it. No secrets.", "Jargon: contract version, runtime state, dependency truth"),
"admin-planning-lookups|actions.reactivated": ("\"{kind}.{key}\" turned back on. It appears in planning lists after the next load.", "Jargon: governed selects"),
"admin-planning-lookups|editor.advanced": ("Advanced settings (JSON)", "Jargon: raw metadata override"),
"admin-risk-models|composer.payload": ("Model settings. Factors must add up to 1.00. Bands run 0 to 100.", "Jargon: payload; middot separators"),
"dashboard|national.coverage.example": ("{count} completed inspections. No annual target is set.", "Jargon: governed"),
"factories-cr|cr.openViolationsUnavailable": ("Not available. Violations have no open or closed status yet.", "Jargon: runtime, governed state"),
"factories-cr|licenses.degraded": ("Licence details are unavailable. Registration details still show.", "Jargon: degraded, CR identity"),
"factories-cr|section.degraded": ("This section is unavailable. Other sections still work.", "Jargon: source section, degraded"),
"factories-cr|source.noSla": ("Freshness is shown as recorded. No staleness limit is set.", "Jargon: source fact, unapproved threshold, inferred"),
"field-my-tasks|list.online": ("Online. Showing your assignments.", "Jargon: server-scoped"),
"planning|bulk.title": ("Plan bulk visits", "Drops criteria and targeting from a title"),
"reviews|detail.cmp.unavailMetadata": ("Section order not compared.", "Jargon: metadata"),
"visits|detail.unverifiedManual": ("Entered manually, not yet verified", "Jargon: pending reconciliation"),
"admin-access|governance.heading": ("How access is controlled", "Jargon: governance, surface"),
"admin-access|manage.selfTarget": ("This is your own account. Another administrator must change your access.", "24 words to 11; drops self-escalation guard"),
"admin-dashboard-config|context": ("KPI catalogue and settings. Filtered to your access. Two people required to publish.", "Jargon: governed, maker-checker"),
"admin-dashboard-config|drafts.footer": ("Writing and publishing must be done by two different administrators.", "27 words to 10; removes phantom role names"),
"admin-delegation|subtitle": ("Temporarily give another user your authority.", "Jargon: governed role, authorized user"),
"admin-devices|result.invalid_request": ("Select a command and give a reason.", "Jargon: governed, meaningful"),
"admin-enforcement-recommendations|error.backend_guard_required": ("Recording a decision is not available yet.", "Drops internal database detail"),
"admin-enforcement-recommendations|permissionsUnavailable.body": ("Your permissions could not be checked. Reload the page.", "18 words to 9"),
"admin-enforcement-recommendations|readonly.body": ("You can view this list. You need an Admin role to decide.", "Removes phantom roles Operations and Compliance Admin"),
"admin-enforcement-recommendations|scope.body": ("Decisions are read-only for now. When enabled, a decision will only be recorded. It will not apply a penalty or close a factory.", "48 words to 23"),
"admin-factory-data|history.noBatchProvenance": ("No batch history recorded", "Jargon: provenance"),
"admin-factory-data|result.CSV_SCHEMA_VERSION_UNSUPPORTED": ("This file version is not supported.", "Jargon: CSV schema version"),
"admin-gis|banner.body": ("These values are recorded on every check-in. Only the administrator can change official coordinates. Edit each factory's geofence on the map below.", "44 words to 22"),
"admin-gis|spatial.banner.body": ("Only the administrator can move the official factory pin. These are extra working layers.", "Drops geofence and accuracy engine defaults"),
"admin-integrations|governance.note": ("This page reads live data limited to your access. Missing records show as unavailable or empty.", "42 words to 16"),
"admin-integrations|governance.title": ("How integrations are controlled", "Jargon: governance, surface"),
"admin-integrations|metrics.configuredNote": ("Details and address recorded", "Jargon: contract, runtime address"),
"admin-items|audit.body": ("Open Audit on an item to see its history.", "19 words to 9"),
"admin-items|audit.unavailable": ("History could not load. Reload to try again.", "Drops this does not mean empty"),
"admin-items|degradedBody": ("The clause list could not load. Everything else on this page still works.", "Jargon: degraded, catalogue rendered"),
"admin-items|footer": ("Items belong to regulations and are reused across checklists. Turning one off keeps its history. Only administrators can edit.", "31 words to 19; removes phantom roles"),
"admin-items|gov.body": ("Anyone signed in can view this. Only administrators can edit. Turning an item off keeps its history and records a reason. Every change is logged.", "42 words to 25"),
"admin-items|gov.heading": ("How this catalogue is controlled", "Jargon: governed"),
"admin-items|legacyNote": ("These controls still work, but new changes should start as a Compliance Configuration Request.", "36 words to 14"),
"admin-localization|governance.heading": ("How translations are controlled", "Jargon: governance, surface"),
"admin-localization|subtitle": ("English and Arabic text for every screen", "Jargon: source, localized surface"),
"admin-notifications|result.notice_not_configured": ("No provider is set up for this channel. Saved as a test only.", "Jargon: provider not available, live provider"),
"admin-notifications|subtitle": ("Rules for which event goes to which channel and recipient", "Jargon: governed, template, delivery service"),
"admin-packages|governance.heading": ("How packages are controlled", "Jargon: governance, surface"),
"admin-packages|register.immutable": ("Published and locked. Create a new draft to change it.", "19 words to 10"),
"admin-packages|templates.hint": ("Versioned templates that packages and penalty mappings can use.", "Jargon: governed"),
"admin-planning-lookups|actions.invalidKind": ("Choose a lookup type. Nothing was changed.", "Jargon: governed"),
"admin-planning-lookups|actions.jsonInvalid": ("The advanced settings are not valid JSON. Nothing was changed.", "Jargon: raw metadata override"),
"admin-planning-lookups|actions.jsonNotObject": ("The advanced settings must be a JSON object. Nothing was changed.", "Jargon: raw metadata override"),
"admin-planning-lookups|actions.keyExistsInsert": ("\"{key}\" already exists in \"{kind}\". Edit that row instead.", "Drops nothing was changed as redundant"),
"admin-planning-lookups|actions.keyExistsUpdate": ("\"{key}\" already exists in \"{kind}\". Keys must be unique.", "18 words to 9"),
"admin-planning-lookups|editor.rawJson": ("Advanced settings (JSON)", "Jargon: raw metadata override"),
"admin-planning-lookups|editor.rawJsonHint": ("Optional. Replaces all advanced settings for this row.", "Jargon: non-empty JSON object, metadata"),
"admin-planning-lookups|governance.body": ("These values fill every planning list. Every change records who made it and what changed.", "26 words to 15"),
"admin-planning-lookups|subtitle": ("Reference values used across planning", "Jargon: governed"),
"admin-planning-status|governance.heading": ("Set by workflow configuration", "Jargon: governed"),
"admin-risk|banner.body": ("Weights and bands are live settings. Only the risk owner can change them. Every save is logged.", "42 words to 17"),
}

def walk(o, p=""):
    if isinstance(o, dict):
        for k, v in o.items(): yield from walk(v, f"{p}.{k}" if p else k)
    elif isinstance(o, list):
        for i, v in enumerate(o): yield from walk(v, f"{p}[{i}]")
    elif isinstance(o, str): yield p, o

current = {}
for path in sorted(glob.glob(os.path.join(LOCALES, "*.json"))):
    ns = os.path.basename(path)[:-5]
    for k, v in walk(json.load(open(path, encoding="utf-8"))):
        current[f"{ns}|{k}"] = v

rows, missing = [], []
for ref, (to_be, reason) in R.items():
    if ref not in current:
        missing.append(ref); continue
    ns, key = ref.split("|", 1)
    rows.append([ns, key, role_for(ns), current[ref], to_be, reason])
rows.sort(key=lambda r: (r[2], r[0], r[1]))

if missing:
    print("KEYS NOT FOUND IN LOCALES:", *missing, sep="\n  ")
    sys.exit(1)

AUB, LILAC, HAIR, WAS, WILL = "413259", "7E61AC", "EFEBF5", "FBEEF0", "EAF4EE"
wb = Workbook(); ws = wb.active; ws.title = "Plain English"
ws.sheet_view.showGridLines = False
HEAD = ["Namespace", "Key", "Role", "As-is", "To-be", "Reason"]
ws.append(HEAD)
for r in rows: ws.append(r)
for i in range(1, len(HEAD) + 1):
    c = ws.cell(row=1, column=i)
    c.fill = PatternFill("solid", fgColor=AUB)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 26
last = len(rows) + 1
t = Table(displayName="PlainEnglish", ref=f"A1:{get_column_letter(len(HEAD))}{last}")
t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
ws.add_table(t)
for w, col in zip([20, 34, 12, 62, 62, 40], "ABCDEF"):
    ws.column_dimensions[col].width = w
edge = Side(style="thin", color="D8D3E0")
for r in range(2, last + 1):
    for i in range(1, len(HEAD) + 1):
        c = ws.cell(row=r, column=i)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = Border(bottom=edge)
    ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=WAS)
    ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=WILL)
    ws.cell(row=r, column=3).font = Font(bold=True, color=AUB)
    ws.cell(row=r, column=6).font = Font(color="6B6280")
ws.freeze_panes = "D2"
ws.auto_filter.ref = t.ref

out = os.path.join(ROOT, "outputs", "MIM-Plain-English-Revisions.xlsx")
wb.save(out)
words = lambda s: len(s.split())
before = sum(words(r[3]) for r in rows); after = sum(words(r[4]) for r in rows)
from collections import Counter
print(f"{out}\n{len(rows)} rows · 1 sheet · {len(HEAD)} columns")
print("roles:", dict(sorted(Counter(r[2] for r in rows).items())))
print(f"words: {before} -> {after}  ({round((1-after/before)*100)}% shorter)")
