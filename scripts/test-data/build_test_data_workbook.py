#!/usr/bin/env python3
"""Build the MIM Inspection test-data pack workbook from the governed registers.

Reads every CSV in product-contract/test-data-architecture/registers and emits a
single formatted .xlsx. The CSVs stay the source of truth; the workbook is a
regenerated view, so editing a register and re-running this script is the only
supported way to change the pack.
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[2]
REGISTERS = ROOT / "product-contract" / "test-data-architecture" / "registers"

AUBERGINE = "413259"
FILL_HEAD = PatternFill("solid", fgColor=AUBERGINE)
FILL_NOTE = PatternFill("solid", fgColor="F1EDF6")
FILL_INPUT = PatternFill("solid", fgColor="FFF7E0")
HAIRLINE = Side(style="thin", color="D8D3E0")
BORDER = Border(bottom=HAIRLINE)

SHEETS = [
    ("00 Read me", None, None),
    ("01 Screen data map", "screen-data-map.csv",
     "Every menu item, every panel, every number on it, and the exact tables behind it. "
     "absence_class_today says what the screen shows with an empty database; unlocked_by says which data layer turns it into a real value."),
    ("02 Measure register", "measure-register.csv",
     "All 28 governed KPIs exactly as they exist in apps/web/src/lib/dashboard-kpi/registry.ts. "
     "blocker_class separates the measures test data can fix from the ones needing a migration or a governance ruling."),
    ("03 Blocked measures", "blocked-measures.csv",
     "The eleven measures that are not simply waiting for rows. Read can_test_data_fix_it before promising anything to the training team."),
    ("04 Absence vocabulary", "absence-vocabulary.csv",
     "The five distinct kinds of nothing. Today the platform collapses Zero and Not applicable into Unavailable, which makes an empty database look like a broken product."),
    ("05 Data layers", "layers.csv",
     "Five layers, loaded low to high and unloaded high to low. L0 is never written and never deleted; audit events are never deleted at all."),
    ("06 Journey stops", "journey-stops.csv",
     "The unit of test data is one Inspection Journey. Its terminal stop decides every row it creates. "
     "Terminal counts add up to the cohort; attributes and modifiers overlay on top of journeys that already exist."),
    ("07 Entity volumes", "entity-volumes.csv",
     "Unit-level inventory: what each entity is, how its primary key is derived, its parent, and how many rows each volume profile creates. "
     "Yellow columns are the ones you edit to change a profile."),
    ("08 Personas", "personas.csv",
     "Forty-nine fictitious identities. The two boundary personas exist to prove refusal paths, not to be demonstrated as working users."),
    ("09 Load and unload", "load-unload-runbook.csv",
     "The full operator runbook. Step 6 always runs before step 7. Step 9 is the whole-data unload."),
]

READ_ME = [
    ("MIM Inspection — test data pack", ""),
    ("", ""),
    ("What this workbook is", "The single place where the test data for every screen is proposed, reviewed and signed off before it is loaded."),
    ("Source of truth", "product-contract/test-data-architecture/registers/*.csv — this file is generated from them."),
    ("Regenerate", "python3 scripts/test-data/build_test_data_workbook.py"),
    ("", ""),
    ("How to read it", ""),
    ("1", "Sheet 01 is the map: pick a menu item and you see every number on it and where that number comes from."),
    ("2", "Sheet 04 explains the five kinds of nothing. This is the part the training team must understand first."),
    ("3", "Sheet 06 is the unit. One Inspection Journey is one row of test data; its terminal stop decides everything downstream."),
    ("4", "Sheet 07 is what you edit to change how much data gets created."),
    ("5", "Sheet 09 is how you load it and how you take it all away again."),
    ("", ""),
    ("The one thing to know before the demo", ""),
    ("", "Most of the Unavailable chips on the dashboard are an empty database, not a broken platform."),
    ("", "Sixty-seven of the eighty-two screen rows in sheet 01 become real values the moment test data is loaded."),
    ("", "Eight more need an Administration policy version published — that is also part of the load, layer L1."),
    ("", "Only six are genuinely blocked on a migration or a governance ruling. Those are listed in sheet 03."),
    ("", ""),
    ("Rules this pack does not break", ""),
    ("", "No governed value is invented. Penalty amounts, risk weights, SLAs and cycle targets come from approved configuration or the measure stays blocked."),
    ("", "No real person or establishment is represented. Every name, CR number and coordinate is fictitious."),
    ("", "Audit events are never seeded and never deleted."),
    ("", "The loader refuses any target that is not on the non-production allowlist."),
]

EDITABLE = {"smoke", "demo", "qa", "perf", "demo_count", "qa_count", "demo_target_value", "count_in_demo"}


def style_header(ws, ncols, title, note):
    end = get_column_letter(max(ncols, 1))
    ws.merge_cells(f"A1:{end}1")
    cell = ws["A1"]
    cell.value = title
    cell.fill = FILL_HEAD
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{end}2")
    note_cell = ws["A2"]
    note_cell.value = note
    note_cell.fill = FILL_NOTE
    note_cell.font = Font(italic=True, color="413259")
    note_cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34


def add_sheet(wb, name, csv_name, note):
    rows = list(csv.reader((REGISTERS / csv_name).open(encoding="utf-8")))
    header, body = rows[0], rows[1:]
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    style_header(ws, len(header), name.split(" ", 1)[1], note)
    ws.append([h.replace("_", " ") for h in header])
    for row in body:
        ws.append(row)
    head_row = 3
    for idx in range(1, len(header) + 1):
        cell = ws.cell(row=head_row, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7E61AC")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[head_row].height = 30
    last = head_row + len(body)
    ref = f"A{head_row}:{get_column_letter(len(header))}{last}"
    table = Table(displayName="T" + "".join(ch for ch in name if ch.isalnum()), ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(table)
    for col_idx, key in enumerate(header, start=1):
        letter = get_column_letter(col_idx)
        width = max(len(key), *(len(r[col_idx - 1]) for r in body)) if body else len(key)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 46)
        if key in EDITABLE:
            for row_idx in range(head_row + 1, last + 1):
                ws.cell(row=row_idx, column=col_idx).fill = FILL_INPUT
    for row_idx in range(head_row + 1, last + 1):
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ws.freeze_panes = ws[f"A{head_row + 1}"]
    return len(body)


def add_read_me(wb):
    ws = wb.create_sheet("00 Read me")
    ws.sheet_view.showGridLines = False
    style_header(ws, 2, "Read me", "Start here, then work left to right through the numbered sheets.")
    for left, right in READ_ME:
        ws.append([left, right])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 116
    for row in ws.iter_rows(min_row=3):
        row[0].font = Font(bold=True, color="413259")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "outputs" / "MIM-Inspection-Test-Data-Pack-v1.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    add_read_me(wb)
    total = 0
    for name, csv_name, note in SHEETS:
        if csv_name is None:
            continue
        total += add_sheet(wb, name, csv_name, note)
    wb.save(out)
    print(f"{out} · {len(wb.sheetnames)} sheets · {total} register rows")


if __name__ == "__main__":
    main()
