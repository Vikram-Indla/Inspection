# -*- coding: utf-8 -*-
"""One sheet of decisions to sign off before the cohort is loaded for training.

Every row: what a trainee will see, why it is that way, and an empty column for
your ruling. Nothing here is settled until that column is filled.

Run: python3 scripts/test-data/build_consent_sheet.py
"""
import csv, json, os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "product-contract", "test-data-architecture", "seed-source")
rd = lambda n: list(csv.DictReader(open(os.path.join(SRC, n), encoding="utf-8")))
js = lambda n: json.load(open(os.path.join(SRC, n), encoding="utf-8"))

J, F, P = rd("journeys.csv"), rd("factories.csv"), rd("personas.csv")
POL, IT = rd("policies.csv"), rd("inspection_items.csv")
RP, CP = js("route-preview.json"), js("compliance-preview.json")
MP = js("map-preview.json")
st = Counter(j["planning_status"] for j in J)
vt = Counter(j["visit_type"] for j in J)
ans = sum(int(j["responses"]) for j in J)

# area, what a trainee sees, current value, why it is this way
ROWS = [
 ("Volume", "Total visits across all screens", f"{len(J)}",
  "Enough to fill every tab and every filter without any combination landing empty."),
 ("Volume", "Establishments", f"{len(F)} across 13 regions, 27 industrial sites",
  "Every region carries approved work, so the region lens has no empty row."),
 ("Volume", "Users who can sign in", f"{len(P)}: 5 admin, 6 planner, 8 supervisor, 24 inspector, 2 boundary",
  "The two boundary users prove refusal paths — a multi-role user and one with no access."),
 ("Volume", "Checklist answers", f"{ans:,}",
  "Follows from package size per visit type, not a chosen number."),
 ("Volume", "Months of history", "12, anchored to 22 Aug 2026",
  "24 of 60 approved visits sit inside the rolling 30 days so Analytics has signal."),

 ("Planning", "Draft tab", f"{st['draft']}", "A planner's unpublished work."),
 ("Planning", "Pending supervision", f"{st['pending_supervision']}", "Waiting on a supervisor decision."),
 ("Planning", "Published", f"{st['published']}", "Approved to run — the bulk of the board."),
 ("Planning", "Returned", f"{st['returned']}", "Sent back to the planner."),
 ("Planning", "Cancelled", f"{st['cancelled']}", "Each carries a recorded reason."),
 ("Planning", "Expired", f"{st['expired']}", "Window passed while still unassigned."),

 ("Visit types", "Routine", f"{vt['routine']} visits · 80 checklist items",
  "Fire, civil defence, occupational safety, industrial security."),
 ("Visit types", "Complaint", f"{vt['complaint']} visits · 34 items",
  "Civil defence, environment, food and drug — targeted, not a full sweep."),
 ("Visit types", "Licensing", f"{vt['licensing']} visits · 26 items",
  "Industry ministry, municipal, standards."),
 ("Visit types", "Follow-up", f"{vt['follow_up']} visits · 12 items",
  "Re-checks only what failed last time. This is why answer counts differ by type."),

 ("Operations", "Inspectors in motion on the day", f"{MP['map']['live_journeys']}",
  "The live map has movement rather than one static pin."),
 ("Operations", "GPS events", f"{MP['map']['geo_events']}",
  "Real tracks: departure, telemetry along the route, then arrival."),
 ("Operations", "Geofence overrides", f"{MP['map']['overrides']}",
  "Each carries a distance from site and a written reason — the supervisor conversation."),
 ("Operations", "Inspector scheduling", "5 slots per inspector per day",
  "The platform refuses overlapping assignments. The cohort has to be schedulable."),

 ("Review", "Awaiting decision", f"{RP['reviews']['queue_pending']}", "The supervisor queue on open."),
 ("Review", "Approved / returned / rejected",
  f"{RP['reviews']['approve']} / {RP['reviews']['returned']} / {RP['reviews']['reject']}",
  "Only approved work counts toward the compliance rate."),

 ("Compliance", "Issuing authorities", f"{CP['authorities']}",
  "Real Saudi regulators. An inspector must see the correct authority on a finding."),
 ("Compliance", "Regulations / clauses / items", f"{CP['regulations']} / {CP['clauses']} / {CP['items']}",
  "Clause and item wording is synthetic and marked so. Only the authority names are real."),
 ("Compliance", "Violation codes / penalty mappings",
  f"{CP['violation_codes']} / {CP['penalty_mappings']}", "Penalty amounts are not invented."),
 ("Compliance", "Configuration requests", f"{CP['config_requests']['total']} across 6 states",
  "Fills the Approval Queue at every workflow step."),

 ("Dashboard", "National compliance rate", f"{RP['dashboard']['STR-KPI-001_compliance_rate']}%",
  "Computed, not set. Changing the cohort changes this number."),
 ("Dashboard", "Approval rate", f"{RP['dashboard']['STR-KPI-004_approval_rate']}%", "Computed."),
 ("Dashboard", "Inspection coverage", f"{MP['unlocked_by_policy']['STR-KPI-007_coverage_pct']}% against an 80% target",
  "Below target on purpose — a full green dashboard teaches nothing."),
 ("Dashboard", "Four measures still blank",
  "Risk distribution · Licence exposure · Repeat violation rate · Violation trend",
  "No data fills these. They need application changes, which are NOT proposed here."),

 ("Establishments", "Names", "All fictitious, from local place names per region",
  "The current seed names real companies and attaches risk scores to them."),
 ("Establishments", "CR numbers", "53 new-regime 7-prefixed, 20 legacy Riyadh 1010",
  "Royal Decree M/83 abolished city-linked registers in Apr 2025; both forms circulate until 2030."),
 ("Establishments", "Licence numbers", "10-digit numeric, format UNVERIFIED",
  "No public source documents the real format. A trainee cross-checking one gets no hit."),
 ("Establishments", "Coordinates", "9 published, 4 district-level, 60 governorate-derived",
  "Right region and city, within a few km. Fine for map pins, not for navigating to a plot."),

 ("Policies", "Inspection cycle", "12 months, 80% annual target",
  "TEST VALUE. Unlocks coverage and uninspected-factories. Needs your number."),
 ("Policies", "SLA urgency warning", "at 75% of the window elapsed",
  "TEST VALUE. Unlocks 'expiring soon'. Needs your number."),
 ("Policies", "KPI targets", "85% compliance, 80% approval",
  "TEST VALUE. Draws the target line on two tiles. Needs your number."),

 ("Load and unload", "Load", "One command, ~30 seconds, idempotent",
  "Re-running upserts rather than duplicating. Proved on a real database."),
 ("Load and unload", "unload --confirm", "Removes pre-submission rows only",
  "Respects every immutability trigger. Works without superuser."),
 ("Load and unload", "reset --confirm", "Removes the whole batch, needs superuser",
  "Suspends triggers for one transaction. Refuses a production-looking database."),
 ("Load and unload", "Audit trail after reset", "Kept",
  "The record of what the training run did stays readable. Say if you want it cleared too."),
 ("Load and unload", "Repeatability", "Identical every run — no clock, no random seed",
  "load → reset → load returns the same 82.9%. Verified twice on a real database."),
]

AUB, LIL, HAIR, INPUT = "413259", "7E61AC", "EFEBF5", "FFF7E0"
wb = Workbook(); ws = wb.active; ws.title = "Consent"
ws.sheet_view.showGridLines = False
HEAD = ["Area", "What a trainee sees", "Proposed", "Why it is this way", "Your decision", "Change it to"]
ws.append(HEAD)
for r in ROWS: ws.append(list(r) + ["", ""])
for i in range(1, len(HEAD) + 1):
    c = ws.cell(row=1, column=i)
    c.fill = PatternFill("solid", fgColor=AUB); c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 28
last = len(ROWS) + 1
dv = DataValidation(type="list", formula1='"Approved,Change,Discuss"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"E2:E{last}")
edge = Side(style="thin", color="D8D3E0")
prev = None
for r in range(2, last + 1):
    area = ws.cell(row=r, column=1).value
    for i in range(1, len(HEAD) + 1):
        c = ws.cell(row=r, column=i)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = Border(bottom=edge, top=Side(style="thin", color=LIL) if area != prev else None)
    ws.cell(row=r, column=1).font = Font(bold=True, color=AUB)
    ws.cell(row=r, column=3).font = Font(bold=True, color="241C33")
    ws.cell(row=r, column=4).font = Font(color="6B6280", size=10)
    for col in (5, 6): ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=INPUT)
    prev = area
for w, col in zip([16, 34, 34, 56, 15, 30], "ABCDEF"): ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{last}"

out = os.path.join(ROOT, "outputs", "MIM-Test-Data-Consent.xlsx")
wb.save(out)
print(f"{out}\n{len(ROWS)} decisions · 1 sheet")
print("areas:", dict(Counter(r[0] for r in ROWS)))
