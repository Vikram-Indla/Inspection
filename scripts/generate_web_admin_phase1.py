#!/usr/bin/env python3
"""Generate the Web/Admin Phase 1 planning baseline from approved sources.

This script writes only lightweight Git artifacts. Source workbooks, documents,
design HTML, images, and archives remain external per the documentation policy.
Run with the bundled Codex Python runtime, which provides openpyxl.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


TASK_ID = "TASK-WEB-ADMIN-PHASE1-PLAN-001"
CHANGE_ID = "CC-WEB-ADMIN-PHASE1-001"
BASE_SHA = "6fc27d3f654a79d2aa6ef659b0879b35b9eb5b6d"
ALLOWED_DISPOSITIONS = {
    "PHASE1_WEB",
    "PHASE1_ADMIN",
    "PHASE1_SHARED_BACKEND",
    "PHASE2_IPAD_DEFERRED",
    "EXTERNAL_CONTRACT_BLOCKED",
    "OPEN_BUSINESS_DECISION",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, headers: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def requirement_target(sheet: str, capability: str) -> tuple[str, str, str, str]:
    text = capability.lower()
    if sheet == "Visit Planning-Planning":
        if "bulk" in text:
            return "MOD-05", "M2", "/planning/bulk/**", "SAQEEL Planning;SAQEEL Planning Identity"
        if "single" in text or "immediate" in text:
            return "MOD-04", "M2", "/planning/single;/planning/immediate", "SAQEEL Planning"
        if any(word in text for word in ("assign", "schedule", "inspector", "capacity")):
            return "MOD-06", "M2", "/planning/**;/visits/**", "SAQEEL Planning;SAQEEL Visits"
        return "MOD-03", "M2", "/planning/**;/visits/**", "SAQEEL Planning;SAQEEL Visits"
    if sheet == "Visit Planning- Mangment":
        return "MOD-03", "M2", "/planning/**;/visits/**", "SAQEEL Planning;SAQEEL Visits"
    if sheet == "Inspection Execution - PreStart":
        return "MOD-07", "PHASE2", "/field/**", "PHASE2_IPAD_DEFERRED"
    if sheet == "Inspection excution-Start jou":
        if any(word in text for word in ("evidence", "violation", "finding", "penalty", "action form")):
            return "MOD-12", "PHASE2", "/field/**", "PHASE2_IPAD_DEFERRED"
        if any(word in text for word in ("submit", "offline", "synchron", "signature", "version")):
            return "MOD-13", "PHASE2", "/field/**", "PHASE2_IPAD_DEFERRED"
        if any(word in text for word in ("journey", "arrival", "location", "gps", "geofence")):
            return "MOD-08", "PHASE2", "/field/**", "PHASE2_IPAD_DEFERRED"
        return "MOD-09", "PHASE2", "/field/**", "PHASE2_IPAD_DEFERRED"
    if sheet == "Inspection Excution - Virtual ":
        return "MOD-10", "M11", "/virtual/**", "SAQEEL Virtual"
    if sheet == "Inspection Excution-Level2Flow":
        return "MOD-14", "M5", "/reviews/**", "SAQEEL Review & Approval"
    if sheet == "Factory 360 ":
        return "MOD-15", "M4", "/factories/**", "SAQEEL Factories;SAQEEL Factory 360"
    if sheet == "Ops center ":
        return "MOD-16", "M3", "/operations/**", "SAQEEL Operations Center;SAQEEL Operations Live"
    if sheet == "Complince ":
        return "MOD-17", "M6", "/admin/compliance-requests/**;/admin/regulations/**;/admin/violations;/enforcement", "SAQEEL Compliance Requests;SAQEEL Enforcement"
    raise ValueError(f"Unmapped source sheet: {sheet!r}")


def requirement_disposition(sheet: str, row_text: str) -> tuple[str, str, str]:
    lowered = row_text.lower()
    # Field/iPad-primary customer rows always stay deferred; shared backend
    # dependencies are linked separately and never relabel the source row.
    if sheet in {"Inspection Execution - PreStart", "Inspection excution-Start jou"}:
        deps = []
        for token, dep in (
            ("audit", "AUDIT"), ("transition", "WORKFLOW"), ("submission", "IMMUTABLE_VERSION"),
            ("notification", "NOTIFICATION"), ("synchron", "OFFLINE_CONFLICT"), ("rls", "RLS"),
        ):
            if token in lowered:
                deps.append(dep)
        return "PHASE2_IPAD_DEFERRED", "", ";".join(deps)

    open_rules = (
        (r"ai planning summary|recommend(?:ation)? of inspection priorit|ai recommendation", "DEC-026"),
        (r"risk (?:weight|formula|threshold|model)|health score", "DEC-001;DEC-028"),
        (r"geofence.*(?:threshold|radius|accuracy)|gps accuracy.*threshold", "DEC-002"),
        (r"sla calendar|working calendar|escalation threshold", "DEC-003"),
        (r"digital signature|legal signature|pki", "DEC-009"),
        (r"planning supervisor|supervisor.*approver|approver.*supervisor", "DEC-TGT-001"),
    )
    for pattern, decision in open_rules:
        if re.search(pattern, lowered):
            return "OPEN_BUSINESS_DECISION", decision, ""

    external_rules = (
        (r"industry shared|shared api", "DEC-027"),
        (r"senaei.*(?:submit|write|trigger)|external submission", "DEC-TGT-008"),
        (r"video provider|video integration|live video", "DEC-TGT-006"),
        (r"otp provider", "DEC-007"),
        (r"notification provider|delivery receipt|sms provider|email provider", "DEC-TGT-009"),
    )
    for pattern, decision in external_rules:
        if re.search(pattern, lowered):
            return "EXTERNAL_CONTRACT_BLOCKED", decision, ""

    return ("PHASE1_ADMIN" if sheet == "Complince " else "PHASE1_WEB"), "", ""


def build_requirements(customer_dir: Path, output_dir: Path) -> tuple[list[dict[str, object]], Counter]:
    workbook_path = customer_dir / "Inspection Project(2).xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    fields = [
        "module", "capability", "purpose", "user_flow", "user_actions", "system_actions",
        "ai_actions", "inputs", "outputs", "business_rules", "personas", "dependencies",
        "source_acceptance_criteria",
    ]
    rows: list[dict[str, object]] = []
    sequence = 0
    for worksheet in workbook.worksheets:
        for source_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            normalized = [clean(value) for value in values[:13]]
            if not any(normalized):
                continue
            sequence += 1
            source = dict(zip(fields, normalized))
            original = " | ".join(f"{key}={value}" for key, value in source.items() if value)
            disposition, decision_ids, shared_deps = requirement_disposition(worksheet.title, original)
            module_id, package_id, routes, designs = requirement_target(worksheet.title, source["capability"])
            requirement_id = f"CR-{sequence:03d}"
            rows.append({
                "requirement_id": requirement_id,
                "source_id": "SRC-CUST-001",
                "source_sha256": sha256(workbook_path),
                "source_sheet": worksheet.title,
                "source_coordinate": f"{worksheet.title}!{source_row}",
                **source,
                "original_source_text": original,
                "disposition": disposition,
                "disposition_alias": "DEFERRED_TO_PHASE_2_IPAD" if disposition == "PHASE2_IPAD_DEFERRED" else "",
                "target_module_id": module_id,
                "implementation_package": package_id,
                "target_routes": routes,
                "target_designs": designs,
                "acceptance_ids": f"WA-AC-{sequence:04d}",
                "decision_or_blocker_ids": decision_ids,
                "shared_backend_dependencies": shared_deps,
                "evidence_ids": f"WA-EV-{sequence:04d}",
                "status": "BASELINED_NOT_IMPLEMENTED",
            })
    if sequence != 478:
        raise RuntimeError(f"Expected 478 requirements, found {sequence}")
    headers = list(rows[0].keys())
    write_csv(output_dir / "REQUIREMENT_BASELINE.csv", headers, rows)
    counts = Counter(row["disposition"] for row in rows)
    return rows, counts


def build_source_manifest(customer_dir: Path, output_dir: Path) -> None:
    authority = {
        "Inspection Project(2).xlsx": "A1_CUSTOMER_MANDATORY_478_ROW_BASELINE",
        "dashboard(4).xlsx": "A1_CUSTOMER_COMPANION",
        "Ipad Phas1.xlsx": "A1_CUSTOMER_COMPANION_PHASE2_DEFERRED",
        "Opearation Center(4).xlsx": "A1_CUSTOMER_COMPANION",
        "Planning.docx": "A1_CUSTOMER_DETAILED_SPECIFICATION",
        "Inspection_Execution_Consolidated.docx": "A1_CUSTOMER_DETAILED_SPECIFICATION_WITH_PHASE2_SCOPE",
        "Factory_360_Specification(3).docx": "A1_CUSTOMER_DETAILED_SPECIFICATION",
        "Complince Mang.docx": "A1_CUSTOMER_DETAILED_SPECIFICATION",
        "Inspection_API_documentation.json": "A2_CUSTOMER_API_CONTRACT",
    }
    rows = []
    for index, name in enumerate(authority, start=1):
        path = customer_dir / name
        rows.append({
            "source_id": f"SRC-CUST-{index:03d}",
            "source_name": name,
            "logical_path": f"MIM_Inspection_Web_Admin_Codex_Plan_Starter.zip!/01_APPROVED_CUSTOMER_SOURCES.zip!/{name}",
            "size_bytes": path.stat().st_size,
            "sha256": sha256(path),
            "authority": authority[name],
            "version": "PRODUCT_OWNER_SUPPLIED_2026-07-23",
            "external_storage_pointer": "/Users/vikramindla/Attachment dump/MIM_Inspection_Web_Admin_Codex_Plan_Starter.zip",
            "phase1_use": "DEFERRED_SOURCE_ONLY" if name == "Ipad Phas1.xlsx" else "REQUIREMENTS_OR_CONTRACT_INPUT",
            "git_storage": "PROHIBITED_BINARY_EXTERNAL_ONLY",
            "verified_date": "2026-07-23",
        })
    write_csv(output_dir / "SOURCE_MANIFEST.csv", list(rows[0].keys()), rows)


def route_from_file(app_root: Path, path: Path, suffix: str) -> str:
    parts = list(path.relative_to(app_root).parent.parts)
    parts = [part for part in parts if not (part.startswith("(") and part.endswith(")"))]
    route_parts = []
    for part in parts:
        if part.startswith("[") and part.endswith("]"):
            route_parts.append(f":{part[1:-1]}")
        else:
            route_parts.append(part)
    route = "/" + "/".join(route_parts)
    if suffix == "api" and not route.startswith("/api"):
        route = "/api" + route
    return route or "/"


def owner_for_route(route: str) -> str:
    if route.startswith("/field") or route.startswith("/api/field"):
        return "PHASE2"
    if route == "/dashboard":
        return "M1"
    if route.startswith("/planning") or route.startswith("/visits"):
        return "M2"
    if route.startswith("/operations") or route == "/api/routing/eta":
        return "M3"
    if route.startswith("/factories"):
        return "M4"
    if route.startswith("/reviews") or route in {"/cases", "/tasks"}:
        return "M5"
    if route in {"/enforcement", "/committee"} or route.startswith((
        "/admin/compliance-", "/admin/regulations", "/admin/violations", "/admin/risk",
        "/admin/enforcement-recommendations", "/admin/bulk-violations",
    )):
        return "M6"
    if route.startswith("/reports") or route in {"/evidence-ocr", "/ai/suggestions", "/incident-reports"}:
        return "M7"
    if route.startswith(("/admin/access", "/admin/security-access", "/admin/devices")):
        return "M8"
    if route.startswith(("/admin/integrations", "/admin/gis", "/admin/execution", "/admin/operations")):
        return "M10"
    if route.startswith("/admin"):
        return "M9"
    if route == "/api/shell/search":
        return "F0"
    if route == "/" or route.startswith(("/login", "/reset", "/launch", "/portal", "/profile", "/virtual")):
        return "M11"
    return "UNASSIGNED"


def nearest_file(start: Path, stop: Path, name: str) -> str:
    current = start
    while True:
        candidate = current / name
        if candidate.exists():
            return candidate.as_posix()
        if current == stop or stop not in current.parents:
            return ""
        current = current.parent


def build_routes(repo: Path, output_dir: Path, requirements: list[dict[str, object]]) -> list[dict[str, object]]:
    app_root = repo / "apps/web/src/app"
    e2e_root = repo / "apps/web/e2e"
    tests = {path: path.read_text(encoding="utf-8", errors="replace") for path in e2e_root.glob("*.ts")}
    nav_text = (repo / "apps/web/src/lib/shell-navigation.ts").read_text(encoding="utf-8")
    nav_hrefs = set(re.findall(r'href:\s*"([^"]+)"', nav_text))
    routes: list[dict[str, object]] = []
    candidates = [
        (path, "PAGE") for path in app_root.rglob("page.tsx")
        if not path.relative_to(app_root).as_posix().startswith("reference/")
    ]
    candidates += [(path, "API") for path in (app_root / "api").rglob("route.ts")]
    for path, route_type in candidates:
        route = route_from_file(app_root, path, "page" if route_type == "PAGE" else "api")
        directory = path.parent
        local_files = sorted(directory.glob("*.ts")) + sorted(directory.glob("*.tsx"))
        combined = "\n".join(item.read_text(encoding="utf-8", errors="replace") for item in local_files)
        imports = sorted(set(re.findall(r'from\s+["\'](@/(?:lib|components)/[^"\']+)["\']', combined)))
        tables = sorted(set(re.findall(r'\.from\(\s*["\']([^"\']+)["\']', combined)))
        rpcs = sorted(set(re.findall(r'\.rpc\(\s*["\']([^"\']+)["\']', combined)))
        roles = sorted(set(re.findall(r'\b(?:planner|inspector|reviewer|ops|leadership|compliance_admin|form_admin|workflow_admin|security_admin|gis_admin|risk_owner)\b', combined)))
        states = [state for state in ("loading", "empty", "error", "degraded", "unauthorized", "conflict", "stale", "not-found") if state.replace("-", " ") in combined.lower() or state in combined.lower()]
        static_prefix = route.split(":", 1)[0].rstrip("/")
        route_tests = []
        if route != "/":
            for test_path, test_text in tests.items():
                if route in test_text or (static_prefix and static_prefix in test_text):
                    route_tests.append(test_path.relative_to(repo).as_posix())
        package = owner_for_route(route)
        phase = "PHASE2_IPAD_DEFERRED" if package == "PHASE2" else "PHASE1_WEB_ADMIN"
        routes.append({
            "route": route,
            "route_type": route_type,
            "phase_scope": phase,
            "owner_package": package,
            "requirement_ids": package_requirement_ids(requirements, package),
            "design_authority": package_designs(package),
            "acceptance_ids": f"WA-{package}-AC-001..006" if package != "PHASE2" else "PHASE2_DEFERRED_NO_PHASE1_ACCEPTANCE",
            "dependencies": ";".join(imports) or "ROUTE_LOCAL_DISCOVERY_REQUIRED",
            "page_or_handler_file": path.relative_to(repo).as_posix(),
            "layout_file": nearest_file(directory, app_root, "layout.tsx").replace(repo.as_posix() + "/", ""),
            "loading_file": (directory / "loading.tsx").relative_to(repo).as_posix() if (directory / "loading.tsx").exists() else "",
            "error_file": (directory / "error.tsx").relative_to(repo).as_posix() if (directory / "error.tsx").exists() else "",
            "not_found_file": (directory / "not-found.tsx").relative_to(repo).as_posix() if (directory / "not-found.tsx").exists() else "",
            "navigation_entry": "DIRECT" if route.split("?", 1)[0] in {href.split("?", 1)[0] for href in nav_hrefs} else "DEEP_LINK_OR_ALIAS",
            "breadcrumb": "PRESENT_OR_COMPONENT" if "breadcrumb" in combined.lower() else "TO_VERIFY_IN_RUNTIME",
            "persona_permission_guard": ";".join(roles) or ("AUTHENTICATED_ROUTE_GROUP" if "/(app)/" in path.as_posix() else "ROUTE_LOCAL_OR_PUBLIC"),
            "server_actions_services": ";".join(imports),
            "supabase_tables_views": ";".join(tables),
            "supabase_rpcs": ";".join(rpcs),
            "workflow_audit_versioning": ";".join(word.upper() for word in ("workflow", "audit", "version", "immutable") if word in combined.lower()),
            "required_states_observed": ";".join(states),
            "mock_hardcoded_degraded_review": "REVIEW_REQUIRED" if any(word in combined.lower() for word in ("mock", "stub", "not yet", "unavailable", "degraded")) else "NO_STATIC_MARKER",
            "tests": ";".join(route_tests) or "TEST_GAP_REQUIRES_PACKAGE_FOCUSED_SPEC",
            "evidence_status": "CURRENT_IMPLEMENTATION_EVIDENCE_ONLY",
        })
    routes.sort(key=lambda row: (row["route_type"], row["route"]))
    write_csv(output_dir / "ROUTE_INVENTORY.csv", list(routes[0].keys()), routes)
    return routes


def build_migration_register(output_dir: Path, routes: list[dict[str, object]]) -> list[dict[str, object]]:
    rows = []
    flag_routes = ("/login", "/admin/gis", "/admin/risk", "/admin/integrations", "/virtual", "/operations/live")
    for route in (item for item in routes if item["route_type"] == "PAGE" and item["phase_scope"] == "PHASE1_WEB_ADMIN"):
        uncertain = (
            str(route["route"]).startswith(flag_routes)
            or route["tests"] == "TEST_GAP_REQUIRES_PACKAGE_FOCUSED_SPEC"
            or route["mock_hardcoded_degraded_review"] == "REVIEW_REQUIRED"
        )
        strategy = "FEATURE_FLAG_OR_GUARDED_PREVIEW_REQUIRED" if uncertain else "DIRECT_REPLACEMENT_AFTER_CERTIFICATION"
        rows.append({
            "migration_id": f"WA-MIG-{len(rows)+1:03d}",
            "current_route": route["route"],
            "current_files": route["page_or_handler_file"],
            "owner_package": route["owner_package"],
            "saqeel_design_reference": route["design_authority"],
            "requirement_ids": route["requirement_ids"],
            "acceptance_ids": route["acceptance_ids"],
            "existing_behavior": f"Preserve current navigation={route['navigation_entry']}; breadcrumb={route['breadcrumb']}; states={route['required_states_observed'] or 'runtime discovery required'}",
            "backend_contracts": f"services={route['server_actions_services'] or 'none statically detected'}; tables/views={route['supabase_tables_views'] or 'none statically detected'}; RPCs={route['supabase_rpcs'] or 'none statically detected'}; workflow/audit/version={route['workflow_audit_versioning'] or 'runtime verification required'}",
            "permissions": route["persona_permission_guard"],
            "current_tests": route["tests"],
            "required_certification": "functional parity;negative paths;RLS/RBAC;audit/versioning;visual diff;Arabic/RTL;accessibility;responsive;performance;cross-module regression",
            "replacement_strategy": strategy,
            "flag_reason": "UNCERTAINTY_OR_TEST_GAP_REQUIRES_CONTROLLED_EXPOSURE" if uncertain else "",
            "cutover_method": "Canonical route cutover only after all acceptance gates pass and Product Owner approval is recorded",
            "rollback_method": "Restore retained pre-replacement implementation or switch server-evaluated flag; no destructive database rollback",
            "legacy_retention": "RETAIN_UNTIL_STABILIZATION_AND_PRODUCT_OWNER_REMOVAL_APPROVAL",
            "completion_status": "PLANNED_NOT_IMPLEMENTED",
        })
    write_csv(output_dir / "CURRENT_TO_TARGET_MIGRATION.csv", list(rows[0].keys()), rows)
    return rows


DESIGN_TARGETS = {
    "SAQEEL AI Studio.dc.html": ("M7", "/ai/suggestions", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Business users"),
    "SAQEEL Admin Detail.dc.html": ("M9", "/admin/regulations/:id;/admin/packages;/admin/items/:id/runtime-preview;/admin/risk;/planning/plans/:id", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Configuration roles"),
    "SAQEEL Admin Execution.dc.html": ("M10", "/admin/execution", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Configuration roles"),
    "SAQEEL Admin Extended.dc.html": ("M9", "/admin/bulk-violations;/admin/enforcement-recommendations;/admin/security-access;/admin/devices", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Admin roles"),
    "SAQEEL Admin Form Builder.dc.html": ("M9", "/admin/packages/:id/designer", "NEW_ROUTE_REQUIRED", "Form Admin"),
    "SAQEEL Admin Geofence.dc.html": ("M10", "/admin/gis", "DESIGN_CONFLICT_REQUIRES_DECISION", "GIS Admin"),
    "SAQEEL Admin Integrations Data.dc.html": ("M10", "/admin/integrations/factory-data", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Technical Admin"),
    "SAQEEL Admin Integrations.dc.html": ("M10", "/admin/integrations", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Technical Admin"),
    "SAQEEL Admin Lookups copy.dc.html": ("M9", "/admin/localization", "EXACT_REPLICATION", "Configuration roles"),
    "SAQEEL Admin Lookups.dc.html": ("M9", "/admin/localization", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Configuration roles"),
    "SAQEEL Admin Role Override.dc.html": ("M8", "/admin/access?view=roles", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Security Admin"),
    "SAQEEL Admin SENAI Data.dc.html": ("M10", "/admin/integrations/senaei", "NEW_ROUTE_REQUIRED", "Technical Admin"),
    "SAQEEL Admin Virtual Premium.dc.html": ("M10", "/admin/virtual", "NEW_ROUTE_REQUIRED", "Technical Admin"),
    "SAQEEL Admin Workflow Builder.dc.html": ("M9", "/admin/workflows", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Workflow Admin"),
    "SAQEEL Admin.dc.html": ("M9", "/admin", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Admin roles"),
    "SAQEEL Auth States.dc.html": ("M11", "/launch;/launch/no-workspace;/reset", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "All users"),
    "SAQEEL Cases.dc.html": ("M5", "/cases", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Case roles"),
    "SAQEEL Committee.dc.html": ("M6", "/committee", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Committee roles"),
    "SAQEEL Compliance Requests.dc.html": ("M6", "/admin/compliance-requests;/admin/compliance-requests/new;/admin/compliance-requests/:id", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Compliance Admin"),
    "SAQEEL Control Panel.dc.html": ("M10", "/admin/operations", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Technical Admin"),
    "SAQEEL Dashboard Config.dc.html": ("M9", "/admin/dashboard-config", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Business Admin"),
    "SAQEEL Design System.dc.html": ("F0", "FOUNDATION_ONLY", "EXACT_REPLICATION", "All personas"),
    "SAQEEL Enforcement.dc.html": ("M6", "/enforcement;/admin/violations", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Compliance roles"),
    "SAQEEL Evidence OCR.dc.html": ("M7", "/evidence-ocr", "REPLICATION_WITH_REQUIRED_MISSING_STATE", "Authorized users"),
    "SAQEEL Executive Overview.html": ("M1", "/dashboard?view=strategic", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Leadership"),
    "SAQEEL Factories.dc.html": ("M4", "/factories", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Business users"),
    "SAQEEL Factory 360.dc.html": ("M4", "/factories/:id;/factories/cr/:id", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Scoped business users"),
    "SAQEEL Feedback.dc.html": ("M11", "/feedback", "NEW_ROUTE_REQUIRED", "External and internal users"),
    "SAQEEL Incident Reports.dc.html": ("M7", "/incident-reports", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Authorized users"),
    "SAQEEL Inspection Report.dc.html": ("M7", "/reports/inspection/:id", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Authorized report readers"),
    "SAQEEL KPI Management.dc.html": ("M9", "/admin/dashboard-config?view=kpis", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Business Admin"),
    "SAQEEL Login.dc.html": ("M11", "/login", "DESIGN_CONFLICT_REQUIRES_DECISION", "Public/authenticated users"),
    "SAQEEL Operations Center.dc.html": ("M3", "/operations", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Operations and leadership"),
    "SAQEEL Operations Live.dc.html": ("M3", "/operations/live", "REPLICATION_WITH_REQUIRED_MISSING_STATE", "Operations"),
    "SAQEEL Planning Identity.dc.html": ("M2", "/planning/bulk", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Planner"),
    "SAQEEL Planning.dc.html": ("M2", "/planning;/planning/bulk;/planning/single;/planning/immediate;/planning/plans", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Planner"),
    "SAQEEL Portal.dc.html": ("M11", "/portal", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "External portal users"),
    "SAQEEL Profile.dc.html": ("M11", "/profile", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Authenticated users"),
    "SAQEEL Review & Approval.dc.html": ("M5", "/reviews;/reviews/:id;/reviews/:id/started", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Reviewer"),
    "SAQEEL Risk.dc.html": ("M6", "/admin/risk;/admin/risk/models", "DESIGN_CONFLICT_REQUIRES_DECISION", "Risk Owner"),
    "SAQEEL States System.dc.html": ("F0", "FOUNDATION_ONLY", "EXACT_REPLICATION", "All personas"),
    "SAQEEL Tasks.dc.html": ("M5", "/tasks", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Task assignees"),
    "SAQEEL Users Roles.dc.html": ("M8", "/admin/access;/admin/security-access", "CURRENT_BACKEND_PRESERVED_NEW_FRONTEND", "Security Admin"),
    "SAQEEL Virtual.dc.html": ("M11", "/virtual;/virtual/:id", "REPLICATION_WITH_REQUIRED_MISSING_STATE", "Virtual participants"),
    "SAQEEL Visits.dc.html": ("M2", "/visits;/visits/:id;/visits/calendar;/visits/map;/visits/workload", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Planner and Operations"),
    "SAQEEL Web Dashboard.dc.html": ("M1", "/dashboard", "REPLICATION_WITH_REAL_DATA_ADAPTATION", "Operations and leadership"),
}


def compact_requirement_ids(ids: list[str]) -> str:
    numbers = sorted(int(value.split("-")[1]) for value in ids)
    if not numbers:
        return "CR-001..CR-478"
    groups: list[str] = []
    start = previous = numbers[0]
    for number in numbers[1:]:
        if number == previous + 1:
            previous = number
            continue
        groups.append(f"CR-{start:03d}" if start == previous else f"CR-{start:03d}..CR-{previous:03d}")
        start = previous = number
    groups.append(f"CR-{start:03d}" if start == previous else f"CR-{start:03d}..CR-{previous:03d}")
    return ";".join(groups)


def package_requirement_ids(requirements: list[dict[str, object]], package: str) -> str:
    if package == "F0" or package == "UNASSIGNED":
        return "CR-001..CR-478"
    source_package = "PHASE2" if package == "PHASE2" else package
    ids = [str(row["requirement_id"]) for row in requirements if row["implementation_package"] == source_package]
    return compact_requirement_ids(ids)


def package_designs(package: str) -> str:
    for package_id, _name, _routes, designs, _dependencies, _order in PACKAGE_DEFINITIONS:
        if package_id == package:
            return designs
    return "PHASE2_DEFERRED_NO_DESIGN_AUTHORITY"


def build_designs(design_dir: Path, repo: Path, output_dir: Path, requirements: list[dict[str, object]]) -> list[dict[str, object]]:
    inventory = list(csv.DictReader((design_dir / "WEB_ADMIN_DESIGN_INVENTORY.csv").open(encoding="utf-8-sig")))
    rows = []
    source_rows = []
    for item in inventory:
        name = item["design_file"]
        if name not in DESIGN_TARGETS:
            raise RuntimeError(f"Design has no target mapping: {name}")
        package, target_routes, delta, persona = DESIGN_TARGETS[name]
        source = (design_dir / name).read_text(encoding="utf-8", errors="replace")
        preview_match = re.search(r'data-props="([^"]+)"', source)
        width, height = 1440, 900
        viewport_source = "REPOSITORY_APPROVED_DEFAULT"
        if preview_match:
            try:
                props = json.loads(html.unescape(preview_match.group(1)))
                preview = props.get("$preview") or {}
                width, height = int(preview.get("width", width)), int(preview.get("height", height))
                viewport_source = "DESIGN_DECLARED_PREVIEW"
            except (ValueError, TypeError, json.JSONDecodeError):
                pass
        route_files = []
        for route in target_routes.split(";"):
            route_path = route.split("?", 1)[0]
            if route_path in {"FOUNDATION_ONLY", ""}:
                continue
            normalized = route_path.replace(":id", "[id]")
            candidate = repo / "apps/web/src/app/(app)" / normalized.lstrip("/") / "page.tsx"
            public_candidate = repo / "apps/web/src/app" / normalized.lstrip("/") / "page.tsx"
            if candidate.exists():
                route_files.append(candidate.relative_to(repo).as_posix())
            elif public_candidate.exists():
                route_files.append(public_candidate.relative_to(repo).as_posix())
        duplicate = "SAQEEL Admin Lookups.dc.html" if name == "SAQEEL Admin Lookups copy.dc.html" else ""
        source_rows.append({
            "source_id": f"SRC-DES-{len(source_rows)+1:03d}",
            "design_file": name,
            "logical_path": f"MIM_Inspection_Web_Admin_Codex_Plan_Starter.zip!/02_APPROVED_DESIGN_EXPORTS.zip!/{name}",
            "size_bytes": (design_dir / name).stat().st_size,
            "sha256": item["sha256"],
            "authority": item["design_precedence"],
            "version": "PRODUCT_OWNER_SUPPLIED_2026-07-23",
            "external_storage_pointer": "/Users/vikramindla/Attachment dump/MIM_Inspection_Web_Admin_Codex_Plan_Starter.zip",
            "git_storage": "PROHIBITED_BINARY_EXTERNAL_ONLY",
            "duplicate_of": duplicate,
            "verified_date": "2026-07-23",
        })
        rows.append({
            "design_id": f"WA-DES-{len(rows)+1:03d}",
            "design_file": name,
            "sha256": item["sha256"],
            "unique_payload": "NO_ALIAS" if not duplicate else "IDENTICAL_ALIAS",
            "duplicate_of": duplicate,
            "phase_1_scope": item["phase_1_scope"],
            "owner_package": package,
            "target_routes": target_routes,
            "target_persona": persona,
            "requirement_ids": package_requirement_ids(requirements, package),
            "acceptance_ids": f"WA-{package}-AC-001..006",
            "dependencies": "F0" if package != "F0" else "NONE_FOUNDATION_FIRST",
            "evidence_ids": f"WA-{package}-EV-001..006",
            "current_implementation_files": ";".join(route_files),
            "backend_preservation": "PRESERVE_RLS_RBAC_WORKFLOW_AUDIT_VERSIONING_FAIL_CLOSED_PROVIDERS",
            "delta_class": delta,
            "required_states": "loading;empty;error;degraded;unauthorized;stale/conflict;provider-unavailable as applicable",
            "desktop_rtl": "EN/LTR;AR/RTL;logical properties;keyboard/focus parity",
            "reference_viewport": f"{width}x{height}",
            "viewport_source": viewport_source,
            "responsive_matrix": "1024x768;412x915 or 390x844;320x800 reflow",
            "visual_acceptance": "ZERO_UNAPPROVED_DIFF_FIXED_CHROMIUM_FONT_ENVIRONMENT",
            "functional_acceptance": "REAL_DATA_RLS_ACTIONS_AUDIT_NEGATIVE_STATES",
            "deviation_status": "REQUIRES_PRODUCT_OWNER_APPROVAL",
        })
    write_csv(output_dir / "DESIGN_ROUTE_MAP.csv", list(rows[0].keys()), rows)
    write_csv(output_dir / "DESIGN_SOURCE_MANIFEST.csv", list(source_rows[0].keys()), source_rows)
    return rows


PACKAGE_DEFINITIONS = [
    ("F0", "Shared SAQEEL foundation", "Shared shell;components;tokens;i18n/RTL;reference renderer;visual harness", "SAQEEL Design System.dc.html;SAQEEL States System.dc.html", "", "1"),
    ("M1", "Executive and Business Dashboard", "/dashboard", "SAQEEL Executive Overview.html;SAQEEL Web Dashboard.dc.html", "F0", "2"),
    ("M2", "Planning and Visits", "/planning/**;/visits/**", "SAQEEL Planning.dc.html;SAQEEL Planning Identity.dc.html;SAQEEL Visits.dc.html", "F0", "3"),
    ("M3", "Operations Center", "/operations/**;/api/routing/eta", "SAQEEL Operations Center.dc.html;SAQEEL Operations Live.dc.html", "F0", "4"),
    ("M4", "Factories and Factory 360", "/factories/**", "SAQEEL Factories.dc.html;SAQEEL Factory 360.dc.html", "F0", "5"),
    ("M5", "Review Cases and Tasks", "/reviews/**;/cases;/tasks", "SAQEEL Review & Approval.dc.html;SAQEEL Cases.dc.html;SAQEEL Tasks.dc.html", "F0;M2", "6"),
    ("M6", "Compliance Enforcement Committee and Risk", "/enforcement;/committee;/admin/compliance-*/**;/admin/regulations/**;/admin/violations;/admin/risk/**;/admin/enforcement-recommendations;/admin/bulk-violations", "SAQEEL Compliance Requests.dc.html;SAQEEL Enforcement.dc.html;SAQEEL Committee.dc.html;SAQEEL Risk.dc.html", "F0", "7"),
    ("M7", "Reports OCR AI and Incidents", "/reports/**;/evidence-ocr;/ai/suggestions;/incident-reports", "SAQEEL Inspection Report.dc.html;SAQEEL Evidence OCR.dc.html;SAQEEL AI Studio.dc.html;SAQEEL Incident Reports.dc.html", "F0", "8"),
    ("M8", "Admin Identity Roles and Security", "/admin/access;/admin/security-access;/admin/devices", "SAQEEL Users Roles.dc.html;SAQEEL Admin Role Override.dc.html", "F0", "9"),
    ("M9", "Admin Workflows Forms Lookups and KPI", "/admin excluding M6/M8/M10 ownership", "SAQEEL Admin.dc.html;SAQEEL Admin Detail.dc.html;SAQEEL Admin Extended.dc.html;SAQEEL Admin Form Builder.dc.html;SAQEEL Admin Workflow Builder.dc.html;SAQEEL Admin Lookups.dc.html;SAQEEL KPI Management.dc.html;SAQEEL Dashboard Config.dc.html", "F0;M8", "10"),
    ("M10", "Admin Integrations GIS Execution and Provider Health", "/admin/integrations/**;/admin/gis/**;/admin/execution;/admin/operations;/admin/virtual", "SAQEEL Admin Integrations.dc.html;SAQEEL Admin Integrations Data.dc.html;SAQEEL Admin Geofence.dc.html;SAQEEL Admin SENAI Data.dc.html;SAQEEL Admin Execution.dc.html;SAQEEL Admin Virtual Premium.dc.html;SAQEEL Control Panel.dc.html", "F0;M8", "11"),
    ("M11", "Auth Portal Profile Virtual and Shared States", "/;/login;/reset;/launch/**;/portal;/profile;/virtual/**;/feedback", "SAQEEL Login.dc.html;SAQEEL Auth States.dc.html;SAQEEL Portal.dc.html;SAQEEL Profile.dc.html;SAQEEL Virtual.dc.html;SAQEEL Feedback.dc.html", "F0", "12"),
]


def build_packages(repo: Path, output_dir: Path, planning_dir: Path, routes: list[dict[str, object]], requirements: list[dict[str, object]]) -> None:
    by_owner: dict[str, list[str]] = defaultdict(list)
    for route in routes:
        if route["owner_package"] not in {"PHASE2", "UNASSIGNED"}:
            by_owner[str(route["owner_package"])].append(str(route["page_or_handler_file"]))
    package_rows = []
    prompt_dir = planning_dir / "prompts"
    prompt_dir.mkdir(parents=True, exist_ok=True)
    for package_id, name, route_ownership, designs, dependencies, order in PACKAGE_DEFINITIONS:
        exact_files = sorted(set(by_owner.get(package_id, [])))
        if package_id == "F0":
            exact_files += [
                "apps/web/src/components/saqeel/**",
                "apps/web/src/components/Shell.tsx",
                "apps/web/src/components/ShellClient.tsx",
                "apps/web/src/lib/shell-navigation.ts",
                "apps/web/src/lib/i18n.ts",
                "apps/web/src/app/astryx.css",
            ]
        package_rows.append({
            "package_id": package_id,
            "name": name,
            "merge_order": order,
            "depends_on": dependencies,
            "route_ownership": route_ownership,
            "exact_current_files": ";".join(exact_files),
            "design_files": designs,
            "requirement_scope": package_requirement_ids(requirements, package_id),
            "acceptance_ids": f"WA-{package_id}-AC-001..006",
            "database_service_contracts": "DISCOVERED_CURRENT_CONTRACTS_PLUS_EXPLICIT_GAP_RECORD;NO_REMOTE_DDL_WITHOUT_APPROVAL",
            "do_not_touch": "/field/**;apps/web/src/lib/offline.ts;G11 performance branch;main;unowned package files",
            "test_contract": "typecheck;build;focused functional/negative;RLS/RBAC;EN/AR RTL;light/dark;responsive;a11y;visual diff;protected regression",
            "evidence_output": f"INSPECTION_DOCS_ROOT/07_TEST_EVIDENCE_AND_SCREENSHOTS/web-admin-phase1/{package_id}",
            "collision_risks": "Shared shell/i18n/global CSS/migrations remain F0-owned; request F0 follow-up for shared changes",
            "rollback": "Revert package commit/branch; database remediation is additive forward-only",
            "completion_command": f"npm run typecheck && npm run build && npx playwright test <focused-{package_id.lower()}-specs> && node scripts/validate_web_admin_phase1.mjs",
        })
        prompt = f"# {package_id} — {name}\n\n"
        prompt += f"Task: `{TASK_ID}-{package_id}`\nChange control: `{CHANGE_ID}`\n\n"
        prompt += "## Authority and scope\n\n"
        prompt += f"- Routes: `{route_ownership}`\n- Designs: {designs}\n- Requirements: {package_requirement_ids(requirements, package_id)}\n- Depends on: {dependencies or 'none; foundation package'}\n\n"
        prompt += "## Exact file ownership\n\n"
        prompt += "\n".join(f"- `{item}`" for item in exact_files) or "- New package-specific files only."
        prompt += "\n\n## Implementation contract\n\n"
        prompt += "Replicate the supplied design at its declared viewport using real route data and existing services. Preserve RLS/RBAC, canonical transitions, audit, immutable versions, maker-checker, and fail-closed providers. Implement loading, empty, error, degraded, unauthorized, stale/conflict and provider-unavailable states where applicable. Do not copy design mock values into production. Do not edit `/field/**` or the offline engines.\n\n"
        prompt += "Before editing a route, consume its row in `CURRENT_TO_TARGET_MIGRATION.csv`. Direct replacement is allowed only after certification where the row says so. When behavior, permissions, policy, provider, backend parity, tests, or rollback confidence is uncertain, use a server-evaluated feature flag or guarded preview. Never delete the current implementation; retain it for rollback until stabilization and Product Owner removal approval.\n\n"
        prompt += "## Verification and evidence\n\n"
        prompt += "Run typecheck, production build, focused positive/negative and permission tests, EN/AR RTL, light/dark, responsive 1024/412/390/320 checks, accessibility and zero-unapproved-diff visual comparison. Store binary evidence externally and commit only textual manifests. Do not merge, push, deploy, enable a provider, apply remote DDL, or mutate shared data without separate approval.\n\n"
        prompt += f"Completion command template: `{package_rows[-1]['completion_command']}`\n"
        (prompt_dir / f"{package_id}_IMPLEMENTATION_PROMPT.md").write_text(prompt, encoding="utf-8")
    write_csv(output_dir / "PACKAGE_MANIFEST.csv", list(package_rows[0].keys()), package_rows)

    criteria = []
    categories = [
        ("FUNCTIONAL", "Authorized source-backed route behavior and all required interactions are wired to real services."),
        ("NEGATIVE_SECURITY", "Unauthorized, validation, stale/conflict, provider and RLS/RBAC negative paths fail safely."),
        ("VISUAL", "Reference and implementation match at the fixed design viewport with zero unapproved difference."),
        ("RTL_RESPONSIVE_A11Y", "English/LTR and Arabic/RTL pass desktop, 1024, 412/390 and 320 reflow plus keyboard/accessibility checks."),
        ("REGRESSION", "Protected backend, workflow, audit, version, maker-checker and adjacent-route regression passes."),
        ("EVIDENCE", "Requirement/design/route IDs and external binary evidence manifest are complete and truthful."),
    ]
    for package_id, name, *_ in PACKAGE_DEFINITIONS:
        for index, (category, criterion) in enumerate(categories, start=1):
            criteria.append({
                "acceptance_id": f"WA-{package_id}-AC-{index:03d}",
                "package_id": package_id,
                "category": category,
                "criterion": criterion,
                "priority": "P0" if category in {"FUNCTIONAL", "NEGATIVE_SECURITY", "REGRESSION"} else "P1",
                "proof": "TEST_PLUS_RUNTIME_PLUS_TEXTUAL_EVIDENCE_INDEX",
                "status": "PLANNED_NOT_IMPLEMENTED",
            })
    write_csv(output_dir / "ACCEPTANCE_CRITERIA.csv", list(criteria[0].keys()), criteria)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--customer-dir", required=True, type=Path)
    parser.add_argument("--design-dir", required=True, type=Path)
    parser.add_argument("--repo", required=True, type=Path)
    args = parser.parse_args()
    repo = args.repo.resolve()
    output_dir = repo / "product-contract/web-admin-phase1"
    planning_dir = repo / ".planning-pack/web-admin-phase1"
    requirements, counts = build_requirements(args.customer_dir, output_dir)
    build_source_manifest(args.customer_dir, output_dir)
    routes = build_routes(repo, output_dir, requirements)
    migrations = build_migration_register(output_dir, routes)
    designs = build_designs(args.design_dir, repo, output_dir, requirements)
    build_packages(repo, output_dir, planning_dir, routes, requirements)
    summary = {
        "task_id": TASK_ID,
        "change_control": CHANGE_ID,
        "baseline_sha": BASE_SHA,
        "requirement_count": len(requirements),
        "requirement_dispositions": {value: counts.get(value, 0) for value in sorted(ALLOWED_DISPOSITIONS)},
        "page_route_count": sum(1 for row in routes if row["route_type"] == "PAGE"),
        "phase1_page_route_count": sum(1 for row in routes if row["route_type"] == "PAGE" and row["phase_scope"] == "PHASE1_WEB_ADMIN"),
        "phase2_page_route_count": sum(1 for row in routes if row["route_type"] == "PAGE" and row["phase_scope"] == "PHASE2_IPAD_DEFERRED"),
        "api_route_count": sum(1 for row in routes if row["route_type"] == "API"),
        "design_file_count": len(designs),
        "unique_design_payload_count": len({row["sha256"] for row in designs}),
        "package_count": len(PACKAGE_DEFINITIONS),
        "migration_row_count": len(migrations),
        "requirement_traceability_coverage": "478/478",
        "phase1_implementation_scope_rows": counts.get("PHASE1_WEB", 0) + counts.get("PHASE1_ADMIN", 0) + counts.get("PHASE1_SHARED_BACKEND", 0),
        "phase2_deferred_rows": counts.get("PHASE2_IPAD_DEFERRED", 0),
        "open_or_blocked_rows": counts.get("OPEN_BUSINESS_DECISION", 0) + counts.get("EXTERNAL_CONTRACT_BLOCKED", 0),
        "generated_date": "2026-07-23",
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "BASELINE_SUMMARY.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
